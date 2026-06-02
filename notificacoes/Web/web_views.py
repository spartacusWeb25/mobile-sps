from django.views.generic import TemplateView, ListView, View
from django.utils import timezone
from django.db.models import Q
from datetime import timedelta, date
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .mixin import DBAndSlugMixin
from Entidades.models import Entidades
from contas_a_pagar.models import Titulospagar
from contas_a_receber.models import Titulosreceber
from contas_a_pagar.models import Bapatitulos
from contas_a_receber.models import Baretitulos
from Pedidos.models import PedidoVenda
from Orcamentos.models import Orcamentos
from Licencas.models import Empresas, Filiais
from logging import getLogger
from django.db.utils import OperationalError, ProgrammingError


logger = getLogger(__name__)


def _local_today():
    now = timezone.now()
    if timezone.is_naive(now):
        return now.date()
    return timezone.localtime(now).date()


def _week_range(base):
    start = base - timedelta(days=base.weekday())
    end = start + timedelta(days=6)
    return start, end


def _day_bounds(day):
    start = day
    end = day + timedelta(days=1)
    return start, end


class NotificacoesDashboardView(DBAndSlugMixin, TemplateView):
    template_name = 'Notificacoes/dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = _local_today()
        w_start, w_end = _week_range(today)
        d_start, d_end = _day_bounds(today)
        
        # Buscar informações de empresa e filial
        empresa_info = None
        filial_info = None
        try:
            empresa_info = Empresas.objects.using(self.db_alias).filter(empr_codi=self.empresa_id).first()
            filial_info = Filiais.objects.using(self.db_alias).filter(empr_empr=self.empresa_id, empr_codi=self.filial_id).first()
        except Exception:
            pass
        
        fornecedores = dict(
            Entidades.objects.using(self.db_alias)
            .filter(enti_empr=self.empresa_id)
            .values_list('enti_clie', 'enti_nome')
        )

        # Mostrar totais de todas as empresas e filiais
        pagar_all_qs = Titulospagar.objects.using(self.db_alias).all()
        receber_all_qs = Titulosreceber.objects.using(self.db_alias).all()
        
        
        pagas_qs = Bapatitulos.objects.using(self.db_alias).filter(bapa_dpag__gte=d_start, bapa_dpag__lt=d_end)
        recebidas_qs = Baretitulos.objects.using(self.db_alias).filter(bare_dpag__gte=d_start, bare_dpag__lt=d_end)
        
        
        orc_qs = Orcamentos.objects.using(self.db_alias).defer("pedi_stat").filter(pedi_data__gte=d_start, pedi_data__lt=d_end)
        ped_qs = PedidoVenda.objects.using(self.db_alias).all()

        def _to_row(obj, tipo):
            if tipo == 'Paga':
                titu = getattr(obj, 'bapa_titu', '') or getattr(obj, 'titu_titu', '')
                parc = getattr(obj, 'bapa_parc', '') or getattr(obj, 'titu_parc', '')
                valo = getattr(obj, 'bapa_valo_pago', '') or getattr(obj, 'bapa_valo', '') or getattr(obj, 'titu_valo', '')
                forn_id = getattr(obj, 'bapa_forn', None) or getattr(obj, 'titu_forn', None)
                venc = getattr(obj, 'bapa_dpag', '') or getattr(obj, 'titu_venc', '')
                aber = 'T'
            elif tipo == 'Recebida':
                titu = getattr(obj, 'bare_titu', '') or getattr(obj, 'titu_titu', '')
                parc = getattr(obj, 'bare_parc', '') or getattr(obj, 'titu_parc', '')
                valo = getattr(obj, 'bare_valo_pago', '') or getattr(obj, 'bare_valo', '') or getattr(obj, 'titu_valo', '')
                forn_id = getattr(obj, 'bare_clie', None) or getattr(obj, 'titu_clie', None) or getattr(obj, 'titu_forn', None)
                venc = getattr(obj, 'bare_dpag', '') or getattr(obj, 'titu_venc', '')
                aber = 'T'
            else:
                titu = getattr(obj, 'titu_titu', '')
                parc = getattr(obj, 'titu_parc', '')
                valo = getattr(obj, 'titu_valo', '')
                forn_id = getattr(obj, 'titu_forn', None) or getattr(obj, 'titu_clie', None)
                venc = getattr(obj, 'titu_venc', '')
                aber = getattr(obj, 'titu_aber', '')
            return {
                'tipo': tipo,
                'titu_titu': titu,
                'titu_parc': parc,
                'titu_valo': valo,
                'titu_forn': forn_id,
                'forncedor_nome': fornecedores.get(forn_id, ''),
                'titu_venc': venc,
                'titu_aber': aber,
                'empresa_id': self.empresa_id,
                'empresa_nome': getattr(empresa_info, 'empr_nome', '') if empresa_info else '',
                'filial_id': self.filial_id,
                'filial_nome': getattr(filial_info, 'empr_nome', '') if filial_info else '',
            }
        receber_rows = [_to_row(o, 'Receber') for o in receber_all_qs.filter(titu_emis__gte=d_start, titu_emis__lt=d_end)[:10]]
        pagar_rows = [_to_row(o, 'Pagar') for o in pagar_all_qs.filter(titu_emis__gte=d_start, titu_emis__lt=d_end)[:10]]
        pagas_rows = [_to_row(o, 'Paga') for o in pagas_qs.filter(bapa_dpag__gte=d_start, bapa_dpag__lt=d_end)[:10]]
        recebidas_rows = [_to_row(o, 'Recebida') for o in recebidas_qs.filter(bare_dpag__gte=d_start, bare_dpag__lt=d_end)[:10]]
        titulos_criados_hoje = receber_rows + pagar_rows + pagas_rows + recebidas_rows
        w_end_inclusive = w_end + timedelta(days=1)
        pagar_hoje = pagar_all_qs.filter(titu_venc__gte=d_start, titu_venc__lt=d_end)
        pagar_semana = pagar_all_qs.filter(titu_venc__gte=w_start, titu_venc__lt=w_end_inclusive)
        receber_hoje = receber_all_qs.filter(titu_venc__gte=d_start, titu_venc__lt=d_end)
        receber_semana = receber_all_qs.filter(titu_venc__gte=w_start, titu_venc__lt=w_end_inclusive)
        pagas_hoje_count = Bapatitulos.objects.using(self.db_alias).filter(bapa_dpag__gte=d_start, bapa_dpag__lt=d_end).count()
        pagas_semana_count = Bapatitulos.objects.using(self.db_alias).filter(bapa_dpag__gte=w_start, bapa_dpag__lt=w_end_inclusive).count()
        recebidas_hoje_count = Baretitulos.objects.using(self.db_alias).filter(bare_dpag__gte=d_start, bare_dpag__lt=d_end).count()
        recebidas_semana_count = Baretitulos.objects.using(self.db_alias).filter(bare_dpag__gte=w_start, bare_dpag__lt=w_end_inclusive).count()
        orcamentos_hoje = orc_qs.filter(pedi_data__gte=d_start, pedi_data__lt=d_end)
        pedidos_hoje = ped_qs.filter(pedi_data__gte=d_start, pedi_data__lt=d_end)

        pedidos_pisos_hoje_count = None
        try:
            from Pisos.models import Pedidospisos
            pedidos_pisos_hoje_count = (
                Pedidospisos.objects.using(self.db_alias)
                .filter(pedi_data__gte=d_start, pedi_data__lt=d_end)
                .count()
            )
        except (OperationalError, ProgrammingError):
            pedidos_pisos_hoje_count = None
        except Exception:
            pedidos_pisos_hoje_count = None

        ctx.update({
            'today': today,
            'pagar_hoje_count': pagar_hoje.count(),
            'pagar_semana_count': pagar_semana.count(),
            'receber_hoje_count': receber_hoje.count(),
            'receber_semana_count': receber_semana.count(),
            'pagas_hoje_count': pagas_hoje_count,
            'pagas_semana_count': pagas_semana_count,
            'recebidas_hoje_count': recebidas_hoje_count,
            'recebidas_semana_count': recebidas_semana_count,
            'titulos_criados_hoje': titulos_criados_hoje,
            'pagas_rows': pagas_rows,
            'recebidas_rows': recebidas_rows,
            'orcamentos_hoje_count': orcamentos_hoje.count(),
            'pedidos_hoje_count': pedidos_hoje.count(),
            'pedidos_pisos_hoje_count': pedidos_pisos_hoje_count,
            'empresa_id': self.empresa_id,
            'empresa_nome': getattr(empresa_info, 'empr_nome', '') if empresa_info else '',
            'filial_id': self.filial_id,
            'filial_nome': getattr(filial_info, 'empr_nome', '') if filial_info else '',
        })
        return ctx


class TitulosCriadosHojeListView(DBAndSlugMixin, ListView):
    template_name = 'Notificacoes/list_titulos.html'
    paginate_by = 50

    def get_queryset(self):
        today = _local_today()
        tipo = (self.request.GET.get('tipo') or '').lower()
        d_start, d_end = _day_bounds(today)
        
        # Buscar informações de empresa e filial
        empresa_info = None
        filial_info = None
        try:
            empresa_info = Empresas.objects.using(self.db_alias).filter(empr_codi=self.empresa_id).first()
            filial_info = Filiais.objects.using(self.db_alias).filter(empr_empr=self.empresa_id, empr_codi=self.filial_id).first()
        except Exception:
            pass
        
        pagar_qs = Titulospagar.objects.using(self.db_alias).filter(titu_empr=self.empresa_id, titu_fili=self.filial_id, titu_emis__gte=d_start, titu_emis__lt=d_end).order_by('titu_venc', 'titu_titu')
        receber_qs = Titulosreceber.objects.using(self.db_alias).filter(titu_empr=self.empresa_id, titu_fili=self.filial_id, titu_emis__gte=d_start, titu_emis__lt=d_end).order_by('titu_venc', 'titu_titu')
        pagas_qs = Bapatitulos.objects.using(self.db_alias).filter(bapa_empr=self.empresa_id, bapa_fili=self.filial_id, bapa_dpag__gte=d_start, bapa_dpag__lt=d_end).order_by('bapa_dpag', 'bapa_titu')
        recebidas_qs = Baretitulos.objects.using(self.db_alias).filter(bare_empr=self.empresa_id, bare_fili=self.filial_id, bare_dpag__gte=d_start, bare_dpag__lt=d_end).order_by('bare_dpag', 'bare_titu')

        fornecedores = dict(
            Entidades.objects.using(self.db_alias)
            .filter(enti_empr=self.empresa_id)
            .values_list('enti_clie', 'enti_nome')
        )

        def _to_row(obj, tipo_label):
            forn_id = getattr(obj, 'titu_forn', None) or getattr(obj, 'bapa_forn', None) or getattr(obj, 'bare_clie', None)
            return {
                'tipo': tipo_label,
                'titu_titu': getattr(obj, 'titu_titu', '') or getattr(obj, 'bapa_titu', '') or getattr(obj, 'bare_titu', ''),
                'titu_parc': getattr(obj, 'titu_parc', '') or getattr(obj, 'bapa_parc', '') or getattr(obj, 'bare_parc', ''),
                'titu_valo': getattr(obj, 'titu_valo', '') or getattr(obj, 'bapa_valo', '') or getattr(obj, 'bare_valo', ''),
                'titu_forn': forn_id,
                'forncedor_nome': fornecedores.get(forn_id, ''),
                'titu_venc': getattr(obj, 'titu_venc', '') or getattr(obj, 'bapa_dpag', '') or getattr(obj, 'bare_dpag', ''),
                'titu_aber': getattr(obj, 'titu_aber', '') or 'T' if tipo_label in ['Paga','Recebida'] else getattr(obj, 'titu_aber', ''),
                'empresa_id': self.empresa_id,
                'empresa_nome': getattr(empresa_info, 'empr_nome', '') if empresa_info else '',
                'filial_id': self.filial_id,
                'filial_nome': getattr(filial_info, 'empr_nome', '') if filial_info else '',
            }
        if tipo == 'pagar':
            return [_to_row(o, 'Pagar') for o in pagar_qs]
        if tipo == 'receber':
            return [_to_row(o, 'Receber') for o in receber_qs]
        if tipo == 'paga':
            return [_to_row(o, 'Paga') for o in pagas_qs]
        if tipo == 'recebida':
            return [_to_row(o, 'Recebida') for o in recebidas_qs]
        return ([_to_row(o, 'Pagar') for o in pagar_qs] + [_to_row(o, 'Receber') for o in receber_qs] + [_to_row(o, 'Paga') for o in pagas_qs] + [_to_row(o, 'Recebida') for o in recebidas_qs])

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # Buscar informações de empresa e filial
        empresa_info = None
        filial_info = None
        try:
            empresa_info = Empresas.objects.using(self.db_alias).filter(empr_codi=self.empresa_id).first()
            filial_info = Filiais.objects.using(self.db_alias).filter(empr_empr=self.empresa_id, empr_codi=self.filial_id).first()
        except Exception:
            pass
        ctx.update({
            'today': _local_today(),
            'tipo': (self.request.GET.get('tipo') or '').lower(),
            'slug': self.slug,
            'empresa_id': self.empresa_id,
            'empresa_nome': getattr(empresa_info, 'empr_nome', '') if empresa_info else '',
            'filial_id': self.filial_id,
            'filial_nome': getattr(filial_info, 'empr_nome', '') if filial_info else '',
        })
        return ctx


class TitulosAPagarListView(DBAndSlugMixin, ListView):
    model = Titulospagar
    template_name = 'Notificacoes/list_pagar.html'
    paginate_by = 50

    def get_queryset(self):
        today = _local_today()
        period = (self.request.GET.get('period') or 'hoje').lower()
        status = self.request.GET.get('status')
        filter_empresa_id = self.request.GET.get('empresa_id')
        filter_filial_id = self.request.GET.get('filial_id')
        
        # Buscar informações de empresa e filial
        empresa_info = None
        filial_info = None
        try:
            empresa_info = Empresas.objects.using(self.db_alias).filter(empr_codi=self.empresa_id).first()
            filial_info = Filiais.objects.using(self.db_alias).filter(empr_empr=self.empresa_id, empr_codi=self.filial_id).first()
        except Exception:
            pass
        
        # Use filtered empresa/filial if provided, otherwise include ALL (None)
        empresa_id = int(filter_empresa_id) if filter_empresa_id else None
        filial_id = int(filter_filial_id) if filter_filial_id else None

        if status == 'quitado':
            w_start, w_end = _week_range(today)
            d_start, d_end = _day_bounds(today)
            w_end_inclusive = w_end + timedelta(days=1)
            qs = Bapatitulos.objects.using(self.db_alias).all()
            # apply filters only when provided
            if empresa_id:
                qs = qs.filter(bapa_empr=empresa_id)
            if filial_id:
                qs = qs.filter(bapa_fili=filial_id)
            if period == 'semana':
                qs = qs.filter(bapa_dpag__gte=w_start, bapa_dpag__lt=w_end_inclusive)
            else:
                qs = qs.filter(bapa_dpag__gte=d_start, bapa_dpag__lt=d_end)
            qs = qs.order_by('bapa_dpag', 'bapa_titu')
        else:
            qs = Titulospagar.objects.using(self.db_alias).all()
            # apply filters only when provided
            if empresa_id:
                qs = qs.filter(titu_empr=empresa_id)
            if filial_id:
                qs = qs.filter(titu_fili=filial_id)
            qs = qs.filter(
                (Q(titu_emis__isnull=True) | Q(titu_emis__gte=date(1900,1,1))),
                (Q(titu_venc__isnull=True) | Q(titu_venc__gte=date(1900,1,1))),
            )
            if period == 'semana':
                w_start, w_end = _week_range(today)
                qs = qs.filter(titu_venc__gte=w_start, titu_venc__lt=w_end + timedelta(days=1))
            else:
                d_start, d_end = _day_bounds(today)
                qs = qs.filter(titu_venc__gte=d_start, titu_venc__lt=d_end)
            if status == 'aberto':
                qs = qs.filter(titu_aber='A')
            qs = qs.order_by('titu_forn', 'titu_venc', 'titu_titu')
        fornecedores = dict(
            Entidades.objects.using(self.db_alias)
            .filter(enti_empr=self.empresa_id)
            .values_list('enti_clie', 'enti_nome')
        )
        def _to_row(o):
            import re
            forn = getattr(o, 'titu_forn', '') or getattr(o, 'bapa_forn', '')
            valo = getattr(o, 'titu_valo', '') or getattr(o, 'bapa_valo', '') or getattr(o, 'bapa_valo_pago', '')
            parc = getattr(o, 'titu_parc', '') or getattr(o, 'bapa_parc', '')
            venc = getattr(o, 'titu_venc', '') or getattr(o, 'bapa_venc', '')
            dpag = getattr(o, 'bapa_dpag', '')
            aber = 'T' if status == 'quitado' else getattr(o, 'titu_aber', '')
            
            # Get empresa and filial info for this specific row
            row_empresa_id = getattr(o, 'titu_empr', '') or getattr(o, 'bapa_empr', '')
            row_filial_id = getattr(o, 'titu_fili', '') or getattr(o, 'bapa_fili', '')
            
            row_empresa_info = None
            row_filial_info = None
            try:
                if row_empresa_id:
                    row_empresa_info = Empresas.objects.using(self.db_alias).filter(empr_codi=row_empresa_id).first()
                if row_filial_id and row_empresa_info:
                    row_filial_info = Filiais.objects.using(self.db_alias).filter(empr_empr=row_empresa_id, empr_codi=row_filial_id).first()
            except Exception:
                pass
            
            # Prepare display names: filial should show only fantasia when available; empresa trimmed up to 'Ltda'
            filia_display = ''
            if row_filial_info:
                filia_display = getattr(row_filial_info, 'empr_fant', '') or getattr(row_filial_info, 'empr_nome', '')
            empresa_display = ''
            if row_empresa_info:
                raw = getattr(row_empresa_info, 'empr_nome', '')
                m = re.search(r"\bltda\b", raw, flags=re.IGNORECASE)
                if m:
                    empresa_display = raw[:m.end()].strip()
                else:
                    empresa_display = raw
            
            return {
                'titu_titu': getattr(o, 'titu_titu', '') or getattr(o, 'bapa_titu', ''),
                'titu_forn': forn,
                'fornecedor_nome': fornecedores.get(forn, ''),
                'titu_valo': valo,
                'titu_parc': parc,
                'titu_venc': venc,
                'bapa_dpag': dpag,
                'titu_aber': aber,
                'empresa_id': row_empresa_id,
                'empresa_nome': empresa_display,
                'empresa_fantasia': getattr(row_empresa_info, 'empr_fant', '') if row_empresa_info else '',
                'filial_id': row_filial_id,
                'filial_nome': filia_display,
                'filial_fantasia': getattr(row_filial_info, 'empr_fant', '') if row_filial_info else '',
            }
        return [
            _to_row(o) for o in qs
        ]

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        objs = ctx.get('object_list') or getattr(self, 'object_list', [])
        try:
            total_count = len(objs)
            total_sum = sum(float((o.get('titu_valo') if isinstance(o, dict) else getattr(o, 'titu_valo', 0)) or 0) for o in objs)
        except Exception:
            total_count, total_sum = 0, 0.0
        
        # Calculate totals by company
        empresas_totals = {}
        for item in objs:
            emp_id = item.get('empresa_id') if isinstance(item, dict) else getattr(item, 'empresa_id', '')
            emp_nome = item.get('empresa_nome') if isinstance(item, dict) else getattr(item, 'empresa_nome', '')
            valor = float((item.get('titu_valo') if isinstance(item, dict) else getattr(item, 'titu_valo', 0)) or 0)
            
            if emp_id not in empresas_totals:
                empresas_totals[emp_id] = {
                    'nome': emp_nome,
                    'valor': 0.0,
                    'count': 0
                }
            empresas_totals[emp_id]['valor'] += valor
            empresas_totals[emp_id]['count'] += 1
        
        # Buscar informações de empresa e filial
        empresa_info = None
        filial_info = None
        try:
            empresa_info = Empresas.objects.using(self.db_alias).filter(empr_codi=self.empresa_id).first()
            filial_info = Filiais.objects.using(self.db_alias).filter(empr_empr=self.empresa_id, empr_codi=self.filial_id).first()
        except Exception:
            pass
        ctx.update({
            'today': _local_today(),
            'period': (self.request.GET.get('period') or 'hoje').lower(),
            'slug': self.slug,
            'total_count': total_count,
            'total_sum': total_sum,
            'empresas_totals': empresas_totals,
            'empresa_id': self.empresa_id,
            'empresa_nome': getattr(empresa_info, 'empr_nome', '') if empresa_info else '',
            'filial_id': self.filial_id,
            'filial_nome': getattr(filial_info, 'empr_nome', '') if filial_info else '',
        })
        return ctx


class TitulosAReceberListView(DBAndSlugMixin, ListView):
    model = Titulosreceber
    template_name = 'Notificacoes/list_receber.html'
    paginate_by = 50

    def get_queryset(self):
        today = _local_today()
        period = (self.request.GET.get('period') or 'hoje').lower()
        status = self.request.GET.get('status')
        filter_empresa_id = self.request.GET.get('empresa_id')
        filter_filial_id = self.request.GET.get('filial_id')
        
        # Buscar informações de empresa e filial
        empresa_info = None
        filial_info = None
        try:
            empresa_info = Empresas.objects.using(self.db_alias).filter(empr_codi=self.empresa_id).first()
            filial_info = Filiais.objects.using(self.db_alias).filter(empr_empr=self.empresa_id, empr_codi=self.filial_id).first()
        except Exception:
            pass
        
        # Use filtered empresa/filial if provided, otherwise use current
        empresa_id = int(filter_empresa_id) if filter_empresa_id else self.empresa_id
        filial_id = int(filter_filial_id) if filter_filial_id else self.filial_id
        
        if status == 'quitado':
            w_start, w_end = _week_range(today)
            d_start, d_end = _day_bounds(today)
            w_end_inclusive = w_end + timedelta(days=1)
            qs = Baretitulos.objects.using(self.db_alias).filter(bare_empr=empresa_id, bare_fili=filial_id)
            if period == 'semana':
                qs = qs.filter(bare_dpag__gte=w_start, bare_dpag__lt=w_end_inclusive)
            else:
                qs = qs.filter(bare_dpag__gte=d_start, bare_dpag__lt=d_end)
            qs = qs.order_by('bare_dpag', 'bare_titu')
        else:
            qs = Titulosreceber.objects.using(self.db_alias).filter(titu_empr=empresa_id, titu_fili=filial_id)
            qs = qs.filter(
                (Q(titu_emis__isnull=True) | Q(titu_emis__gte=date(1900,1,1))),
                (Q(titu_venc__isnull=True) | Q(titu_venc__gte=date(1900,1,1))),
            )
            if period == 'semana':
                w_start, w_end = _week_range(today)
                qs = qs.filter(titu_venc__gte=w_start, titu_venc__lt=w_end + timedelta(days=1))
            else:
                d_start, d_end = _day_bounds(today)
                qs = qs.filter(titu_venc__gte=d_start, titu_venc__lt=d_end)
            if status == 'aberto':
                qs = qs.filter(titu_aber='A')
            if status == 'quitado':
                qs = qs.filter(titu_aber='T')
            qs = qs.order_by('titu_clie', 'titu_venc', 'titu_titu')
        clientes = dict(
            Entidades.objects.using(self.db_alias)
            .filter(enti_empr=self.empresa_id)
            .values_list('enti_clie', 'enti_nome')
        )
        def _to_row(o):
            clie = getattr(o, 'titu_clie', '') or getattr(o, 'bare_clie', '')
            valo = getattr(o, 'titu_valo', '') or getattr(o, 'bare_valo', '') or getattr(o, 'bare_valo_pago', '')
            parc = getattr(o, 'titu_parc', '') or getattr(o, 'bare_parc', '')
            venc = getattr(o, 'titu_venc', '') or getattr(o, 'bare_venc', '')
            dpag = getattr(o, 'bare_dpag', '')
            aber = 'T' if status == 'quitado' else getattr(o, 'titu_aber', '')
            
            # Get empresa and filial info for this specific row
            row_empresa_id = getattr(o, 'titu_empr', '') or getattr(o, 'bare_empr', '')
            row_filial_id = getattr(o, 'titu_fili', '') or getattr(o, 'bare_fili', '')
            
            row_empresa_info = None
            row_filial_info = None
            try:
                if row_empresa_id:
                    row_empresa_info = Empresas.objects.using(self.db_alias).filter(empr_codi=row_empresa_id).first()
                if row_filial_id and row_empresa_info:
                    row_filial_info = Filiais.objects.using(self.db_alias).filter(empr_empr=row_empresa_id, empr_codi=row_filial_id).first()
            except Exception:
                pass
            
            return {
                'titu_titu': getattr(o, 'titu_titu', '') or getattr(o, 'bare_titu', ''),
                'titu_clie': clie,
                'cliente_nome': clientes.get(clie, ''),
                'titu_valo': valo,
                'titu_parc': parc,
                'titu_venc': venc,
                'bapa_dpag': dpag,
                'titu_aber': aber,
                'empresa_id': row_empresa_id,
                'empresa_nome': getattr(row_empresa_info, 'empr_nome', '') if row_empresa_info else '',
                'empresa_fantasia': getattr(row_empresa_info, 'empr_fant', '') if row_empresa_info else '',
                'filial_id': row_filial_id,
                'filial_nome': getattr(row_filial_info, 'empr_nome', '') if row_filial_info else '',
                'filial_fantasia': getattr(row_filial_info, 'empr_fant', '') if row_filial_info else '',
            }
        return [
            _to_row(o) for o in qs
        ]


    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        objs = ctx.get('object_list') or getattr(self, 'object_list', [])
        try:
            total_count = len(objs)
            total_sum = sum(float((o.get('titu_valo') if isinstance(o, dict) else getattr(o, 'titu_valo', 0)) or 0) for o in objs)
        except Exception:
            total_count, total_sum = 0, 0.0
        
        # Calculate totals by company
        empresas_totals = {}
        for item in objs:
            emp_id = item.get('empresa_id') if isinstance(item, dict) else getattr(item, 'empresa_id', '')
            emp_nome = item.get('empresa_nome') if isinstance(item, dict) else getattr(item, 'empresa_nome', '')
            valor = float((item.get('titu_valo') if isinstance(item, dict) else getattr(item, 'titu_valo', 0)) or 0)
            
            if emp_id not in empresas_totals:
                empresas_totals[emp_id] = {
                    'nome': emp_nome,
                    'valor': 0.0,
                    'count': 0
                }
            empresas_totals[emp_id]['valor'] += valor
            empresas_totals[emp_id]['count'] += 1
        
        # Buscar informações de empresa e filial
        empresa_info = None
        filial_info = None
        try:
            empresa_info = Empresas.objects.using(self.db_alias).filter(empr_codi=self.empresa_id).first()
            filial_info = Filiais.objects.using(self.db_alias).filter(empr_empr=self.empresa_id, empr_codi=self.filial_id).first()
        except Exception:
            pass
        ctx.update({
            'today': _local_today(),
            'period': (self.request.GET.get('period') or 'hoje').lower(),
            'slug': self.slug,
            'total_count': total_count,
            'total_sum': total_sum,
            'empresas_totals': empresas_totals,
            'empresa_id': self.empresa_id,
            'empresa_nome': getattr(empresa_info, 'empr_nome', '') if empresa_info else '',
            'filial_id': self.filial_id,
            'filial_nome': getattr(filial_info, 'empr_nome', '') if filial_info else '',
        })
        return ctx


# Print Views
class ImprimirPagarView(DBAndSlugMixin, TemplateView):
    template_name = 'Notificacoes/print_pagar.html'
    
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from .services import prepare_pagar_print_data
        
        period = self.request.GET.get('period', 'hoje')
        status = self.request.GET.get('status')
        filter_empresa_id = self.request.GET.get('empresa_id')
        filter_filial_id = self.request.GET.get('filial_id')
        
        # Use empresa_id and filial_id from request if available, otherwise None (to include ALL)
        empresa_id = int(filter_empresa_id) if filter_empresa_id else None
        filial_id = int(filter_filial_id) if filter_filial_id else None
        
        print_data = prepare_pagar_print_data(
            self.db_alias,
            empresa_id,
            filial_id,
            period,
            status
        )
        
        ctx.update(print_data)
        ctx.update({
            'slug': self.slug,
        })
        return ctx


class ImprimirReceberView(DBAndSlugMixin, TemplateView):
    template_name = 'Notificacoes/print_receber.html'
    
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from .services import prepare_receber_print_data
        
        period = self.request.GET.get('period', 'hoje')
        status = self.request.GET.get('status')
        filter_empresa_id = self.request.GET.get('empresa_id')
        filter_filial_id = self.request.GET.get('filial_id')
        
        empresa_id = int(filter_empresa_id) if filter_empresa_id else self.empresa_id
        filial_id = int(filter_filial_id) if filter_filial_id else self.filial_id
        
        print_data = prepare_receber_print_data(
            self.db_alias,
            empresa_id,
            filial_id,
            period,
            status
        )
        
        ctx.update(print_data)
        ctx.update({
            'slug': self.slug,
        })
        return ctx


class ImprimirOrcamentosView(DBAndSlugMixin, TemplateView):
    template_name = 'Notificacoes/print_orcamentos.html'
    
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from .services import prepare_orcamentos_print_data
        
        filter_empresa_id = self.request.GET.get('empresa_id')
        filter_filial_id = self.request.GET.get('filial_id')
        
        empresa_id = int(filter_empresa_id) if filter_empresa_id else self.empresa_id
        filial_id = int(filter_filial_id) if filter_filial_id else self.filial_id
        
        print_data = prepare_orcamentos_print_data(
            self.db_alias,
            empresa_id,
            filial_id
        )
        
        ctx.update(print_data)
        ctx.update({
            'slug': self.slug,
        })
        return ctx


class ImprimirPedidosView(DBAndSlugMixin, TemplateView):
    template_name = 'Notificacoes/print_pedidos.html'
    
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from .services import prepare_pedidos_print_data
        
        filter_empresa_id = self.request.GET.get('empresa_id')
        filter_filial_id = self.request.GET.get('filial_id')
        
        empresa_id = int(filter_empresa_id) if filter_empresa_id else self.empresa_id
        filial_id = int(filter_filial_id) if filter_filial_id else self.filial_id
        
        print_data = prepare_pedidos_print_data(
            self.db_alias,
            empresa_id,
            filial_id
        )
        
        ctx.update(print_data)
        ctx.update({
            'slug': self.slug,
        })
        return ctx


class ImprimirTitulosView(DBAndSlugMixin, TemplateView):
    template_name = 'Notificacoes/print_titulos.html'
    
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from .services import prepare_titulos_print_data
        
        tipo = self.request.GET.get('tipo')
        
        print_data = prepare_titulos_print_data(
            self.db_alias,
            self.empresa_id,
            self.filial_id,
            tipo
        )
        
        ctx.update(print_data)
        ctx.update({
            'slug': self.slug,
        })
        return ctx


class OrcamentosHojeListView(DBAndSlugMixin, ListView):
    model = Orcamentos
    template_name = 'Notificacoes/list_orcamentos.html'
    paginate_by = 50

    def get_queryset(self):
        today = _local_today()
        qs = Orcamentos.objects.using(self.db_alias).filter(pedi_empr=self.empresa_id, pedi_fili=self.filial_id, pedi_data=today)
        return qs.order_by('-pedi_nume')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # Buscar informações de empresa e filial
        empresa_info = None
        filial_info = None
        try:
            empresa_info = Empresas.objects.using(self.db_alias).filter(empr_codi=self.empresa_id).first()
            filial_info = Filiais.objects.using(self.db_alias).filter(empr_empr=self.empresa_id, empr_codi=self.filial_id).first()
        except Exception:
            pass
        ctx.update({
            'today': _local_today(),
            'slug': self.slug,
            'empresa_id': self.empresa_id,
            'empresa_nome': getattr(empresa_info, 'empr_nome', '') if empresa_info else '',
            'filial_id': self.filial_id,
            'filial_nome': getattr(filial_info, 'empr_nome', '') if filial_info else '',
        })
        return ctx


class PedidosHojeListView(DBAndSlugMixin, ListView):
    model = PedidoVenda
    template_name = 'Notificacoes/list_pedidos.html'
    paginate_by = 50

    def get_queryset(self):
        today = _local_today()
        qs = PedidoVenda.objects.using(self.db_alias).filter(pedi_empr=self.empresa_id, pedi_fili=self.filial_id, pedi_data=today)
        return qs.order_by('-pedi_nume')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # Buscar informações de empresa e filial
        empresa_info = None
        filial_info = None
        try:
            empresa_info = Empresas.objects.using(self.db_alias).filter(empr_codi=self.empresa_id).first()
            filial_info = Filiais.objects.using(self.db_alias).filter(empr_empr=self.empresa_id, empr_codi=self.filial_id).first()
        except Exception:
            pass
        ctx.update({
            'today': _local_today(),
            'slug': self.slug,
            'empresa_id': self.empresa_id,
            'empresa_nome': getattr(empresa_info, 'empr_nome', '') if empresa_info else '',
            'filial_id': self.filial_id,
            'filial_nome': getattr(filial_info, 'empr_nome', '') if filial_info else '',
        })
        return ctx


class ExportarTitulosCriadosHojeView(DBAndSlugMixin, View):
    def get(self, request, *args, **kwargs):
        today = _local_today()
        d_start, d_end = _day_bounds(today)
        tipo = (request.GET.get('tipo') or '').lower()

        # Buscar informações de empresa e filial
        empresa_info = None
        filial_info = None
        try:
            empresa_info = Empresas.objects.using(self.db_alias).filter(empr_codi=self.empresa_id).first()
            filial_info = Filiais.objects.using(self.db_alias).filter(empr_empr=self.empresa_id, empr_codi=self.filial_id).first()
        except Exception:
            pass

        fornecedores = dict(
            Entidades.objects.using(self.db_alias)
            .filter(enti_empr=self.empresa_id)
            .values_list('enti_clie', 'enti_nome')
        )

        pagar_qs = Titulospagar.objects.using(self.db_alias).filter(titu_empr=self.empresa_id, titu_fili=self.filial_id, titu_emis__gte=d_start, titu_emis__lt=d_end).order_by('titu_venc', 'titu_titu')
        receber_qs = Titulosreceber.objects.using(self.db_alias).filter(titu_empr=self.empresa_id, titu_fili=self.filial_id, titu_emis__gte=d_start, titu_emis__lt=d_end).order_by('titu_venc', 'titu_titu')
        pagas_qs = Bapatitulos.objects.using(self.db_alias).filter(bapa_empr=self.empresa_id, bapa_fili=self.filial_id, bapa_dpag__gte=d_start, bapa_dpag__lt=d_end).order_by('bapa_dpag', 'bapa_titu')
        recebidas_qs = Baretitulos.objects.using(self.db_alias).filter(bare_empr=self.empresa_id, bare_fili=self.filial_id, bare_dpag__gte=d_start, bare_dpag__lt=d_end).order_by('bare_dpag', 'bare_titu')

        def _to_row(obj, tipo):
            if tipo == 'Paga':
                titu = getattr(obj, 'bapa_titu', '') or getattr(obj, 'titu_titu', '')
                parc = getattr(obj, 'bapa_parc', '') or getattr(obj, 'titu_parc', '')
                valo = getattr(obj, 'bapa_valo_pago', '') or getattr(obj, 'bapa_valo', '') or getattr(obj, 'titu_valo', '')
                forn_id = getattr(obj, 'bapa_forn', None) or getattr(obj, 'titu_forn', None)
                venc = getattr(obj, 'bapa_dpag', '') or getattr(obj, 'titu_venc', '')
                aber = 'T'
            elif tipo == 'Recebida':
                titu = getattr(obj, 'bare_titu', '') or getattr(obj, 'titu_titu', '')
                parc = getattr(obj, 'bare_parc', '') or getattr(obj, 'titu_parc', '')
                valo = getattr(obj, 'bare_valo_pago', '') or getattr(obj, 'bare_valo', '') or getattr(obj, 'titu_valo', '')
                forn_id = getattr(obj, 'bare_clie', None) or getattr(obj, 'titu_clie', None) or getattr(obj, 'titu_forn', None)
                venc = getattr(obj, 'bare_dpag', '') or getattr(obj, 'titu_venc', '')
                aber = 'T'
            else:
                titu = getattr(obj, 'titu_titu', '')
                parc = getattr(obj, 'titu_parc', '')
                valo = getattr(obj, 'titu_valo', '')
                forn_id = getattr(obj, 'titu_forn', None) or getattr(obj, 'titu_clie', None)
                venc = getattr(obj, 'titu_venc', '')
                aber = getattr(obj, 'titu_aber', '')
            return {
                'tipo': tipo,
                'titu_titu': titu,
                'titu_parc': parc,
                'titu_valo': valo,
                'titu_forn': forn_id,
                'forncedor_nome': fornecedores.get(forn_id, ''),
                'titu_venc': venc,
                'titu_aber': aber,
                'empresa_id': self.empresa_id,
                'empresa_nome': getattr(empresa_info, 'empr_nome', '') if empresa_info else '',
                'filial_id': self.filial_id,
                'filial_nome': getattr(filial_info, 'empr_nome', '') if filial_info else '',
            }

        rows = []
        if tipo == 'pagar':
            rows = [_to_row(o, 'Pagar') for o in pagar_qs]
        elif tipo == 'receber':
            rows = [_to_row(o, 'Receber') for o in receber_qs]
        elif tipo == 'paga':
            rows = [_to_row(o, 'Paga') for o in pagas_qs]
        elif tipo == 'recebida':
            rows = [_to_row(o, 'Recebida') for o in recebidas_qs]
        else:
            rows = [_to_row(o, 'Pagar') for o in pagar_qs] + [_to_row(o, 'Receber') for o in receber_qs] + [_to_row(o, 'Paga') for o in pagas_qs] + [_to_row(o, 'Recebida') for o in recebidas_qs]

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=titulos_criados_hoje.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = 'Títulos Criados Hoje'

        headers = ['Tipo', 'Título', 'Parcela', 'Valor', 'Fornecedor', 'Nome Fornecedor', 'Vencimento', 'Estado', 'Empresa ID', 'Empresa Nome', 'Filial ID', 'Filial Nome']
        ws.append(headers)

        header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for row in rows:
            ws.append([
                row['tipo'],
                row['titu_titu'],
                row['titu_parc'],
                float(row['titu_valo'] or 0),
                row['titu_forn'],
                row['forncedor_nome'],
                row['titu_venc'],
                'Concluído' if row['titu_aber'] == 'T' else 'Aberto',
                row['empresa_id'],
                row['empresa_nome'],
                row['filial_id'],
                row['filial_nome'],
            ])

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(response)
        return response


class ExportarTitulosAPagarListView(DBAndSlugMixin,View):
    def get(self, request, *args, **kwargs):

        today = _local_today()
        period = (request.GET.get('period') or 'hoje').lower()
        status = request.GET.get('status')

        # Buscar informações de empresa e filial
        empresa_info = None
        filial_info = None
        try:
            empresa_info = Empresas.objects.using(self.db_alias).filter(empr_codi=self.empresa_id).first()
            filial_info = Filiais.objects.using(self.db_alias).filter(empr_empr=self.empresa_id, empr_codi=self.filial_id).first()
        except Exception:
            pass

        if status == 'quitado':
            w_start, w_end = _week_range(today)
            d_start, d_end = _day_bounds(today)
            w_end_inclusive = w_end + timedelta(days=1)
            qs = Bapatitulos.objects.using(self.db_alias).filter(bapa_empr=self.empresa_id, bapa_fili=self.filial_id)
            if period == 'semana':
                qs = qs.filter(bapa_dpag__gte=w_start, bapa_dpag__lt=w_end_inclusive)
            else:
                qs = qs.filter(bapa_dpag__gte=d_start, bapa_dpag__lt=d_end)
            qs = qs.order_by('bapa_dpag', 'bapa_titu')
        else:
            qs = Titulospagar.objects.using(self.db_alias).filter(titu_empr=self.empresa_id, titu_fili=self.filial_id)
            qs = qs.filter(
                (Q(titu_emis__isnull=True) | Q(titu_emis__gte=date(1900,1,1))),
                (Q(titu_venc__isnull=True) | Q(titu_venc__gte=date(1900,1,1))),
            )
            if period == 'semana':
                w_start, w_end = _week_range(today)
                qs = qs.filter(titu_venc__gte=w_start, titu_venc__lt=w_end + timedelta(days=1))
            else:
                d_start, d_end = _day_bounds(today)
                qs = qs.filter(titu_venc__gte=d_start, titu_venc__lt=d_end)
            if status == 'aberto':
                qs = qs.filter(titu_aber='A')
            qs = qs.order_by('titu_forn', 'titu_venc', 'titu_titu').only('titu_titu','titu_parc','titu_venc','titu_aber','titu_forn','titu_valo')

        fornecedores = dict(
            Entidades.objects.using(self.db_alias)
            .filter(enti_empr=self.empresa_id)
            .values_list('enti_clie', 'enti_nome')
        )

        def _to_row(o):
            forn = getattr(o, 'titu_forn', '') or getattr(o, 'bapa_forn', '')
            valo = getattr(o, 'titu_valo', '') or getattr(o, 'bapa_valo', '') or getattr(o, 'bapa_valo_pago', '')
            parc = getattr(o, 'titu_parc', '') or getattr(o, 'bapa_parc', '')
            venc = getattr(o, 'titu_venc', '') or getattr(o, 'titu_venc', '')
            dpag = getattr(o, 'bapa_dpag', '')
            aber = 'T' if status == 'quitado' else getattr(o, 'titu_aber', '')
            return {
                'titu_titu': getattr(o, 'titu_titu', '') or getattr(o, 'bapa_titu', ''),
                'titu_forn': forn,
                'fornecedor_nome': fornecedores.get(forn, ''),
                'titu_valo': valo,
                'titu_parc': parc,
                'titu_venc': venc,
                'bapa_dpag': dpag,
                'titu_aber': aber,
                'empresa_id': self.empresa_id,
                'empresa_nome': getattr(empresa_info, 'empr_nome', '') if empresa_info else '',
                'filial_id': self.filial_id,
                'filial_nome': getattr(filial_info, 'empr_nome', '') if filial_info else '',
            }

        rows = [_to_row(o) for o in qs]

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=contas_pagar.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = 'Contas a Pagar'

        headers = ['Título', 'Fornecedor', 'Nome Fornecedor', 'Valor', 'Parcela', 'Vencimento', 'Data Pagamento', 'Estado', 'Empresa ID', 'Empresa Nome', 'Filial ID', 'Filial Nome']
        ws.append(headers)

        header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for row in rows:
            ws.append([
                row['titu_titu'],
                row['titu_forn'],
                row['fornecedor_nome'],
                float(row['titu_valo'] or 0),
                row['titu_parc'],
                row['titu_venc'],
                row['bapa_dpag'],
                'Quitado' if row['titu_aber'] == 'T' else 'Aberto',
                row['empresa_id'],
                row['empresa_nome'],
                row['filial_id'],
                row['filial_nome'],
            ])

        total_sum = sum(float(row['titu_valo'] or 0) for row in rows)
        total_count = len(rows)

        ws.append([])
        ws.append(['TOTAIS', '', '', total_sum, '', '', '', f'Títulos: {total_count}', '', '', '', ''])

        total_fill = PatternFill(fill_type='solid', fgColor='4472C4')
        total_font = Font(color='FFFFFF', bold=True)
        for cell in ws[ws.max_row]:
            cell.fill = total_fill
            cell.font = total_font

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(response)
        return response


class ExportarTitulosAReceberListView(DBAndSlugMixin,View):
    def get(self, request, *args, **kwargs):

        today = _local_today()
        period = (request.GET.get('period') or 'hoje').lower()
        status = request.GET.get('status')

        # Buscar informações de empresa e filial
        empresa_info = None
        filial_info = None
        try:
            empresa_info = Empresas.objects.using(self.db_alias).filter(empr_codi=self.empresa_id).first()
            filial_info = Filiais.objects.using(self.db_alias).filter(empr_empr=self.empresa_id, empr_codi=self.filial_id).first()
        except Exception:
            pass

        if status == 'quitado':
            w_start, w_end = _week_range(today)
            d_start, d_end = _day_bounds(today)
            w_end_inclusive = w_end + timedelta(days=1)
            qs = Baretitulos.objects.using(self.db_alias).filter(bare_empr=self.empresa_id, bare_fili=self.filial_id)
            if period == 'semana':
                qs = qs.filter(bare_dpag__gte=w_start, bare_dpag__lt=w_end_inclusive)
            else:
                qs = qs.filter(bare_dpag__gte=d_start, bare_dpag__lt=d_end)
            qs = qs.order_by('bare_dpag', 'bare_titu')
        else:
            qs = Titulosreceber.objects.using(self.db_alias).filter(titu_empr=self.empresa_id, titu_fili=self.filial_id)
            qs = qs.filter(
                (Q(titu_emis__isnull=True) | Q(titu_emis__gte=date(1900,1,1))),
                (Q(titu_venc__isnull=True) | Q(titu_venc__gte=date(1900,1,1))),
            )
            if period == 'semana':
                w_start, w_end = _week_range(today)
                qs = qs.filter(titu_venc__gte=w_start, titu_venc__lt=w_end + timedelta(days=1))
            else:
                d_start, d_end = _day_bounds(today)
                qs = qs.filter(titu_venc__gte=d_start, titu_venc__lt=d_end)
            if status == 'aberto':
                qs = qs.filter(titu_aber='A')
            if status == 'quitado':
                qs = qs.filter(titu_aber='T')
            qs = qs.order_by('titu_clie', 'titu_venc', 'titu_titu').only('titu_titu','titu_parc','titu_venc','titu_aber','titu_clie','titu_valo')

        clientes = dict(
            Entidades.objects.using(self.db_alias)
            .filter(enti_empr=self.empresa_id)
            .values_list('enti_clie', 'enti_nome')
        )

        def _to_row(o):
            clie = getattr(o, 'titu_clie', '') or getattr(o, 'bare_clie', '')
            valo = getattr(o, 'titu_valo', '') or getattr(o, 'bare_valo', '') or getattr(o, 'bare_valo_pago', '')
            parc = getattr(o, 'titu_parc', '') or getattr(o, 'bare_parc', '')
            venc = getattr(o, 'titu_venc', '') or getattr(o, 'bare_venc', '')
            dpag = getattr(o, 'bare_dpag', '')
            aber = 'T' if status == 'quitado' else getattr(o, 'titu_aber', '')
            return {
                'titu_titu': getattr(o, 'titu_titu', '') or getattr(o, 'bare_titu', ''),
                'titu_clie': clie,
                'cliente_nome': clientes.get(clie, ''),
                'titu_valo': valo,
                'titu_parc': parc,
                'titu_venc': venc,
                'bapa_dpag': dpag,
                'titu_aber': aber,
                'empresa_id': self.empresa_id,
                'empresa_nome': getattr(empresa_info, 'empr_nome', '') if empresa_info else '',
                'filial_id': self.filial_id,
                'filial_nome': getattr(filial_info, 'empr_nome', '') if filial_info else '',
            }

        rows = [_to_row(o) for o in qs]

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=contas_receber.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = 'Contas a Receber'

        headers = ['Título', 'Cliente', 'Nome Cliente', 'Valor', 'Parcela', 'Vencimento', 'Data Recebimento', 'Estado', 'Empresa ID', 'Empresa Nome', 'Filial ID', 'Filial Nome']
        ws.append(headers)

        header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for row in rows:
            ws.append([
                row['titu_titu'],
                row['titu_clie'],
                row['cliente_nome'],
                float(row['titu_valo'] or 0),
                row['titu_parc'],
                row['titu_venc'],
                row['bapa_dpag'],
                'Recebido' if row['titu_aber'] == 'T' else 'Em Aberto',
                row['empresa_id'],
                row['empresa_nome'],
                row['filial_id'],
                row['filial_nome'],
            ])

        total_sum = sum(float(row['titu_valo'] or 0) for row in rows)
        total_count = len(rows)

        ws.append([])
        ws.append(['TOTAIS', '', '', total_sum, '', '', '', f'Títulos: {total_count}', '', '', '', ''])

        total_fill = PatternFill(fill_type='solid', fgColor='4472C4')
        total_font = Font(color='FFFFFF', bold=True)
        for cell in ws[ws.max_row]:
            cell.fill = total_fill
            cell.font = total_font

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(response)
        return response


class ExportarOrcamentosHojeListView(DBAndSlugMixin,View):
    def get(self, request, *args, **kwargs):

        today = _local_today()
        qs = Orcamentos.objects.using(self.db_alias).filter(pedi_empr=self.empresa_id, pedi_fili=self.filial_id, pedi_data=today)
        rows = qs.order_by('-pedi_nume')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=orcamentos_hoje.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = 'Orçamentos Hoje'

        headers = ['Número', 'Data', 'Cliente', 'Vendedor', 'Status', 'Total']
        ws.append(headers)

        header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for o in rows:
            ws.append([
                o.pedi_nume or "",
                o.pedi_data.strftime("%d/%m/%Y") if o.pedi_data else "",
                o.pedi_clie or "",
                o.pedi_vend or "",
                getattr(o, 'pedi_stat', ''),
                float(o.pedi_tota or 0),
            ])

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(response)
        return response


class ExportarPedidosHojeListView(DBAndSlugMixin,View):
    def get(self, request, *args, **kwargs):

        today = _local_today()
        qs = PedidoVenda.objects.using(self.db_alias).filter(pedi_empr=self.empresa_id, pedi_fili=self.filial_id, pedi_data=today)
        rows = qs.order_by('-pedi_nume')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=pedidos_hoje.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = 'Pedidos Hoje'

        headers = ['Número', 'Data', 'Cliente', 'Vendedor', 'Status', 'Total']
        ws.append(headers)

        header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for p in rows:
            ws.append([
                p.pedi_nume or "",
                p.pedi_data.strftime("%d/%m/%Y") if p.pedi_data else "",
                p.pedi_clie or "",
                p.pedi_vend or "",
                getattr(p, 'pedi_stat', ''),
                float(p.pedi_tota or 0),
            ])

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(response)
        return response


class NotificacoesPorEmpresaFilialView(DBAndSlugMixin, TemplateView):
    template_name = 'Notificacoes/resumo_empresas_filiais.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = _local_today()
        d_start, d_end = _day_bounds(today)
        w_start, w_end = _week_range(today)
        w_end_inclusive = w_end + timedelta(days=1)
        
        try:
            empresas = Empresas.objects.using(self.db_alias).all().order_by('empr_nome')
            filiais = Filiais.objects.using(self.db_alias).all().order_by('empr_empr', 'empr_codi')
        except Exception as e:
            logger.error(f"Error fetching empresas/filiais: {e}")
            empresas = []
            filiais = []
        
        resumo = []
        total_geral = {
            'pagar_hoje': 0,
            'pagar_semana': 0,
            'receber_hoje': 0,
            'receber_semana': 0,
            'pagas_hoje': 0,
            'pagas_semana': 0,
            'recebidas_hoje': 0,
            'recebidas_semana': 0,
            'orcamentos_hoje': 0,
            'pedidos_hoje': 0,
        }
        
        for empresa in empresas:
            filiais_da_empresa = [f for f in filiais if f.empr_empr == empresa.empr_codi]
            
            for filial in filiais_da_empresa:
                try:
                    # Contas a Pagar
                    pagar_hoje = Titulospagar.objects.using(self.db_alias).filter(
                        titu_empr=empresa.empr_codi, 
                        titu_fili=filial.empr_codi,
                        titu_venc__gte=d_start, 
                        titu_venc__lt=d_end
                    ).count()
                    
                    pagar_semana = Titulospagar.objects.using(self.db_alias).filter(
                        titu_empr=empresa.empr_codi, 
                        titu_fili=filial.empr_codi,
                        titu_venc__gte=w_start, 
                        titu_venc__lt=w_end_inclusive
                    ).count()
                    
                    # Contas a Receber
                    receber_hoje = Titulosreceber.objects.using(self.db_alias).filter(
                        titu_empr=empresa.empr_codi, 
                        titu_fili=filial.empr_codi,
                        titu_venc__gte=d_start, 
                        titu_venc__lt=d_end
                    ).count()
                    
                    receber_semana = Titulosreceber.objects.using(self.db_alias).filter(
                        titu_empr=empresa.empr_codi, 
                        titu_fili=filial.empr_codi,
                        titu_venc__gte=w_start, 
                        titu_venc__lt=w_end_inclusive
                    ).count()
                    
                    # Contas Pagas
                    pagas_hoje = Bapatitulos.objects.using(self.db_alias).filter(
                        bapa_empr=empresa.empr_codi, 
                        bapa_fili=filial.empr_codi,
                        bapa_dpag__gte=d_start, 
                        bapa_dpag__lt=d_end
                    ).count()
                    
                    pagas_semana = Bapatitulos.objects.using(self.db_alias).filter(
                        bapa_empr=empresa.empr_codi, 
                        bapa_fili=filial.empr_codi,
                        bapa_dpag__gte=w_start, 
                        bapa_dpag__lt=w_end_inclusive
                    ).count()
                    
                    # Contas Recebidas
                    recebidas_hoje = Baretitulos.objects.using(self.db_alias).filter(
                        bare_empr=empresa.empr_codi, 
                        bare_fili=filial.empr_codi,
                        bare_dpag__gte=d_start, 
                        bare_dpag__lt=d_end
                    ).count()
                    
                    recebidas_semana = Baretitulos.objects.using(self.db_alias).filter(
                        bare_empr=empresa.empr_codi, 
                        bare_fili=filial.empr_codi,
                        bare_dpag__gte=w_start, 
                        bare_dpag__lt=w_end_inclusive
                    ).count()
                    
                    # Orçamentos
                    orcamentos_hoje = Orcamentos.objects.using(self.db_alias).filter(
                        pedi_empr=empresa.empr_codi, 
                        pedi_fili=filial.empr_codi,
                        pedi_data__gte=d_start, 
                        pedi_data__lt=d_end
                    ).count()
                    
                    # Pedidos
                    pedidos_hoje = PedidoVenda.objects.using(self.db_alias).filter(
                        pedi_empr=empresa.empr_codi, 
                        pedi_fili=filial.empr_codi,
                        pedi_data__gte=d_start, 
                        pedi_data__lt=d_end
                    ).count()
                    
                    item = {
                        'empresa_id': empresa.empr_codi,
                        'empresa_nome': empresa.empr_nome,
                        'empresa_fantasia': empresa.empr_fant or '',
                        'filial_id': filial.empr_codi,
                        'filial_nome': filial.empr_nome,
                        'filial_fantasia': filial.empr_fant or '',
                        'tem_logo': bool(filial.empr_logo),
                        'pagar_hoje': pagar_hoje,
                        'pagar_semana': pagar_semana,
                        'receber_hoje': receber_hoje,
                        'receber_semana': receber_semana,
                        'pagas_hoje': pagas_hoje,
                        'pagas_semana': pagas_semana,
                        'recebidas_hoje': recebidas_hoje,
                        'recebidas_semana': recebidas_semana,
                        'orcamentos_hoje': orcamentos_hoje,
                        'pedidos_hoje': pedidos_hoje,
                    }
                    
                    resumo.append(item)
                    
                    # Update total geral
                    total_geral['pagar_hoje'] += pagar_hoje
                    total_geral['pagar_semana'] += pagar_semana
                    total_geral['receber_hoje'] += receber_hoje
                    total_geral['receber_semana'] += receber_semana
                    total_geral['pagas_hoje'] += pagas_hoje
                    total_geral['pagas_semana'] += pagas_semana
                    total_geral['recebidas_hoje'] += recebidas_hoje
                    total_geral['recebidas_semana'] += recebidas_semana
                    total_geral['orcamentos_hoje'] += orcamentos_hoje
                    total_geral['pedidos_hoje'] += pedidos_hoje
                    
                except Exception as e:
                    logger.error(f"Error processing empresa {empresa.empr_codi} filial {filial.empr_codi}: {e}")
                    continue
        
        ctx.update({
            'slug': self.slug,
            'resumo': resumo,
            'total_geral': total_geral,
            'today': today,
        })
        return ctx
