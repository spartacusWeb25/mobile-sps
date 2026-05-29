import json
from decimal import Decimal
from datetime import datetime

from django.views.generic import TemplateView, View
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from core.utils import get_db_from_slug

from Pisos.services.comissao_service import ComissaoService


class ComissaoVendedorView(TemplateView):
    template_name = "Pisos/comissoes_vendedores.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        slug = self.kwargs["slug"]
        db_alias = get_db_from_slug(slug)

        empresa_param = self.request.GET.get("empresa")
        filial_param = self.request.GET.get("filial")

        sess_empresa = (
            self.request.session.get("empresa_id")
            or self.request.session.get("empresa")
            or self.request.session.get("empr_codi")
        )
        sess_filial = (
            self.request.session.get("filial_id")
            or self.request.session.get("filial")
            or self.request.session.get("fili_codi")
        )

        empresa_id = sess_empresa if empresa_param in [None, ""] else empresa_param
        filial_id = sess_filial if filial_param in [None, ""] else filial_param

        try:
            empresa_id = int(empresa_id) if empresa_id not in [None, "", "all"] else None
        except Exception:
            empresa_id = None

        try:
            filial_id = int(filial_id) if filial_id not in [None, "", "all"] else None
        except Exception:
            filial_id = None

        data_inicial = self.request.GET.get("data_inicial")
        data_final = self.request.GET.get("data_final")
        vendedor = self.request.GET.get("vendedor")
        grupo = self.request.GET.get("grupo")
        agrupar_por = self.request.GET.get("agrupar_por") or "vendedor"
        status = self.request.GET.get("status")

        # Set default data_inicial to first day of current month if not provided
        if not data_inicial:
            hoje = datetime.now()
            data_inicial = hoje.replace(day=1).strftime("%Y-%m-%d")

        context["slug"] = slug

        vendedores = list(
            ComissaoService.listar_vendedores(
                db_alias=db_alias,
                empresa_id=empresa_id,
            )
        )
        context["vendedores"] = vendedores

        context["grupos"] = ComissaoService.listar_grupos_comissionados(
            db_alias=db_alias
        )

        try:
            from Licencas.models import Empresas, Filiais

            context["empresas_list"] = list(
                Empresas.objects.using(db_alias).only("empr_codi", "empr_nome").order_by("empr_nome")
            )

            filiais_qs = Filiais.objects.using(db_alias).only("empr_empr", "empr_codi", "empr_nome").order_by("empr_nome")
            if empresa_id is not None:
                filiais_qs = filiais_qs.filter(empr_codi=empresa_id)
            context["filiais_list"] = list(filiais_qs)
        except Exception:
            context["empresas_list"] = []
            context["filiais_list"] = []

        resultado = ComissaoService.calcular_comissoes(
            db_alias=db_alias,
            empresa_id=empresa_id,
            filial_id=filial_id,
            data_inicial=data_inicial,
            data_final=data_final,
            vendedor_codigo=vendedor,
            grupo_codigo=grupo,
            agrupar_por=agrupar_por,
            status=status,
            vendedores=vendedores,
        )
        comissoes = resultado.get("dados", [])
        total_pedidos_periodo = resultado.get("total_pedidos_periodo", 0)
        context["comissoes"] = comissoes

        context["filtros"] = {
            "empresa": "all" if empresa_param == "all" else (str(empresa_id) if empresa_id is not None else ""),
            "filial": "all" if filial_param == "all" else (str(filial_id) if filial_id is not None else ""),
            "data_inicial": data_inicial,
            "data_final": data_final,
            "vendedor": vendedor,
            "grupo": grupo,
            "agrupar_por": agrupar_por,
            "status": status,
        }

        total_vendido = sum((c.get("total_vendido") or 0) for c in comissoes) if comissoes else Decimal("0.00")
        total_comissionado = sum((c.get("total_comissionado") or 0) for c in comissoes) if comissoes else Decimal("0.00")
        total_comissao = sum((c.get("valor_comissao") or 0) for c in comissoes) if comissoes else Decimal("0.00")
        quantidade_pedidos = total_pedidos_periodo
        quantidade_itens = sum((c.get("quantidade_itens") or 0) for c in comissoes)

        context["metricas"] = {
            "total_vendido": total_vendido,
            "total_comissionado": total_comissionado,
            "total_comissao": total_comissao,
            "quantidade_pedidos": quantidade_pedidos,
            "quantidade_itens": quantidade_itens,
        }

        labels = [str(c.get("agrupamento") or "-") for c in comissoes]
        comissoes_vals = [float(c.get("valor_comissao") or 0) for c in comissoes]
        vendas_vals = [float(c.get("total_vendido") or 0) for c in comissoes]

        labels_json = json.dumps(labels, ensure_ascii=False)
        comissoes_json = json.dumps(comissoes_vals)
        vendas_json = json.dumps(vendas_vals)

        context["chart"] = {
            "labels": labels_json,
            "comissoes": comissoes_json,
            "vendas": vendas_json,
        }

        return context


class ExportarComissoesView(View):
    def get(self, request, *args, **kwargs):
        slug = kwargs.get("slug")
        db_alias = get_db_from_slug(slug)

        # Get filters from request
        empresa_param = request.GET.get("empresa")
        filial_param = request.GET.get("filial")

        sess_empresa = (
            request.session.get("empresa_id")
            or request.session.get("empresa")
            or request.session.get("empr_codi")
        )
        sess_filial = (
            request.session.get("filial_id")
            or request.session.get("filial")
            or request.session.get("fili_codi")
        )

        empresa_id = sess_empresa if empresa_param in [None, ""] else empresa_param
        filial_id = sess_filial if filial_param in [None, ""] else filial_param

        try:
            empresa_id = int(empresa_id) if empresa_id not in [None, "", "all"] else None
        except Exception:
            empresa_id = None

        try:
            filial_id = int(filial_id) if filial_id not in [None, "", "all"] else None
        except Exception:
            filial_id = None

        data_inicial = request.GET.get("data_inicial")
        data_final = request.GET.get("data_final")
        vendedor = request.GET.get("vendedor")
        grupo = request.GET.get("grupo")
        agrupar_por = request.GET.get("agrupar_por") or "vendedor"
        status = request.GET.get("status")

        # Handle string "None" values from template
        if data_final == "None":
            data_final = None
        if vendedor == "None":
            vendedor = None
        if grupo == "None":
            grupo = None
        if status == "None":
            status = None

        # Set default data_inicial to first day of current month if not provided
        if not data_inicial:
            hoje = datetime.now()
            data_inicial = hoje.replace(day=1).strftime("%Y-%m-%d")

        # Get vendedores list
        vendedores = list(
            ComissaoService.listar_vendedores(
                db_alias=db_alias,
                empresa_id=empresa_id,
            )
        )

        # Calculate commissions with filters
        resultado = ComissaoService.calcular_comissoes(
            db_alias=db_alias,
            empresa_id=empresa_id,
            filial_id=filial_id,
            data_inicial=data_inicial,
            data_final=data_final,
            vendedor_codigo=vendedor,
            grupo_codigo=grupo,
            agrupar_por=agrupar_por,
            status=status,
            vendedores=vendedores,
        )
        comissoes = resultado.get("dados", [])

        # Create Excel response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=comissoes.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = 'Comissões'

        headers = [
            'Agrupamento',
            'Vendedor',
            'Cód. Vendedor',
            'Grupo/Marca',
            'Origem Comissão',
            'Produto',
            'Cód. Produto',
            'Pedido',
            'Data',
            'Qtd. Itens',
            'Qtd. Pedidos',
            'Total Vendido',
            'Total Comissionado',
            '% Comissão',
            'Valor Comissão'
        ]
        ws.append(headers)

        # Format header
        header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Add data rows
        for c in comissoes:
            vendedor_nome = c.get("vendedor_nome") or "-"
            vendedor_codigo = c.get("vendedor_codigo") or ""

            if c.get("comissao_origem") == "marca":
                grupo_marca = c.get("marca_nome") or "-"
            else:
                grupo_marca = c.get("grupo_descricao") or "-"

            origem_comissao = "Marca" if c.get("comissao_origem") == "marca" else "Grupo"

            produto_nome = c.get("produto_nome") or "-"
            produto_codigo = c.get("produto_codigo") or ""

            pedido_numero = c.get("pedido_numero")
            pedido_str = f"#{pedido_numero}" if pedido_numero else "-"

            # Handle dates - show all dates when multiple exist
            data = c.get("data")
            if data:
                data_str = data.strftime("%d/%m/%Y")
            else:
                # When there's no single date, we need to check if this is from the service
                # The service stores dates in a set but only exports single date
                # For export, we'll show "Múltiplas datas" when grouped by non-date field
                agrupar_por = request.GET.get("agrupar_por") or "vendedor"
                if agrupar_por != "data":
                    data_str = "Múltiplas datas"
                else:
                    data_str = "-"

            ws.append([
                c.get("agrupamento") or "-",
                vendedor_nome,
                vendedor_codigo,
                grupo_marca,
                origem_comissao,
                produto_nome,
                produto_codigo,
                pedido_str,
                data_str,
                c.get("quantidade_itens") or 0,
                c.get("quantidade_pedidos") or 0,
                float(c.get("total_vendido") or 0),
                float(c.get("total_comissionado") or 0),
                float(c.get("percentual") or 0),
                float(c.get("valor_comissao") or 0),
            ])

        # Add totals row
        total_vendido = sum((c.get("total_vendido") or 0) for c in comissoes) if comissoes else Decimal("0.00")
        total_comissionado = sum((c.get("total_comissionado") or 0) for c in comissoes) if comissoes else Decimal("0.00")
        total_comissao = sum((c.get("valor_comissao") or 0) for c in comissoes) if comissoes else Decimal("0.00")
        quantidade_itens = sum((c.get("quantidade_itens") or 0) for c in comissoes)
        quantidade_pedidos = resultado.get("total_pedidos_periodo", 0)

        ws.append([])
        ws.append([
            "TOTAIS",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            quantidade_itens,
            quantidade_pedidos,
            float(total_vendido),
            float(total_comissionado),
            "",
            float(total_comissao),
        ])

        # Format totals row
        total_fill = PatternFill(fill_type='solid', fgColor='4472C4')
        total_font = Font(color='FFFFFF', bold=True)
        for cell in ws[ws.max_row]:
            cell.fill = total_fill
            cell.font = total_font

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(response)
        return response
