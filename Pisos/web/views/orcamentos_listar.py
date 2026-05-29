from datetime import date, datetime
from decimal import Decimal

from django.db.models import Sum, Count
from django.views.generic import ListView, View
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from core.utils import get_db_from_slug
from core.mixins.vendedor_mixin import VendedorEntidadeMixin

from Pisos.models import Orcamentopisos, StatusPisos
from Pisos.services.status_pisos_service import StatusPisosService

from Entidades.models import Entidades


class OrcamentoPisosListView(VendedorEntidadeMixin, ListView):
    model = Orcamentopisos
    template_name = "Pisos/orcamentos_listar.html"
    context_object_name = "orcamentos"
    paginate_by = 50

    def get_queryset(self):
        self.banco = get_db_from_slug(self.kwargs["slug"])
        data_min = date(2020, 1, 1)

        qs = (
            Orcamentopisos.objects
            .using(self.banco)
            .filter(orca_data__gte=data_min)
            .only(
                "orca_empr",
                "orca_fili",
                "orca_nume",
                "orca_clie",
                "orca_vend",
                "orca_data",
                "orca_tota",
                "orca_stat",
            )
            .order_by("-orca_nume")
        )

        qs = self.filter_por_vendedor(qs, "orca_vend")

        numero = self.request.GET.get("orca_nume")
        cliente_nome = self.request.GET.get("cliente_nome")
        vendedor_nome = self.request.GET.get("vendedor_nome")
        status = self.request.GET.get("orca_stat")
        data_inicio = self.request.GET.get("data_inicio")
        data_fim = self.request.GET.get("data_fim")
        empresa_id = self.request.GET.get("empresa_id")
        filial_id = self.request.GET.get("filial_id")

        if numero:
            qs = qs.filter(orca_nume=numero)

        if status not in (None, ""):
            qs = qs.filter(orca_stat=status)

        if data_inicio:
            try:
                di = datetime.strptime(data_inicio, '%Y-%m-%d').date()
                qs = qs.filter(orca_data__gte=di)
            except ValueError:
                pass

        if data_fim:
            try:
                df = datetime.strptime(data_fim, '%Y-%m-%d').date()
                qs = qs.filter(orca_data__lte=df)
            except ValueError:
                pass

        if cliente_nome:
            clientes_ids = Entidades.objects.using(self.banco).filter(
                enti_nome__icontains=cliente_nome
            ).values_list("enti_clie", flat=True)

            qs = qs.filter(orca_clie__in=list(clientes_ids))

        if vendedor_nome:
            # Handle multiple seller names (always use getlist to get all selected values)
            vendedor_nomes = self.request.GET.getlist("vendedor_nome")
            
            # Build Q objects for each seller name with icontains
            from django.db.models import Q
            q_objects = Q()
            for nome in vendedor_nomes:
                if nome:
                    q_objects |= Q(enti_nome__icontains=nome)
            
            vendedores_ids = Entidades.objects.using(self.banco).filter(q_objects).values_list("enti_clie", flat=True)

            qs = qs.filter(orca_vend__in=list(vendedores_ids))

        if empresa_id and empresa_id != "":
            try:
                qs = qs.filter(orca_empr=int(empresa_id))
            except (ValueError, TypeError):
                pass

        if filial_id and filial_id != "":
            try:
                qs = qs.filter(orca_fili=int(filial_id))
            except (ValueError, TypeError):
                pass

        rows = list(qs)

        entidades_ids = set()
        for o in rows:
            if o.orca_clie:
                entidades_ids.add(o.orca_clie)
            if o.orca_vend:
                entidades_ids.add(o.orca_vend)

        entidades = Entidades.objects.using(self.banco).filter(
            enti_clie__in=entidades_ids
        )

        nomes = {
            e.enti_clie: e.enti_nome
            for e in entidades
        }

        if rows:
            empresa = rows[0].orca_empr
            filial = rows[0].orca_fili

            status_map = StatusPisosService.mapa_status(
                banco=self.banco,
                empresa=empresa,
                filial=filial,
                tipo=StatusPisos.TIPO_ORCAMENTO,
            )
        else:
            status_map = {}

        for o in rows:
            status_obj = status_map.get(o.orca_stat)

            o.cliente_nome = nomes.get(o.orca_clie, "")
            o.vendedor_nome = nomes.get(o.orca_vend, "")
            o.status_desc = status_obj.stat_desc if status_obj else "Sem status"
            o.status_cor = status_obj.stat_cor if status_obj else "#6c757d"

        return rows

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        data_min = date(2020, 1, 1)

        # Build filtered queryset for metrics (same as get_queryset)
        qs = (
            Orcamentopisos.objects
            .using(self.banco)
            .filter(orca_data__gte=data_min)
        )
        qs = self.filter_por_vendedor(qs, "orca_vend")

        numero = self.request.GET.get("orca_nume")
        cliente_nome = self.request.GET.get("cliente_nome")
        vendedor_nome = self.request.GET.get("vendedor_nome")
        status = self.request.GET.get("orca_stat")
        data_inicio = self.request.GET.get("data_inicio")
        data_fim = self.request.GET.get("data_fim")
        empresa_id = self.request.GET.get("empresa_id")
        filial_id = self.request.GET.get("filial_id")

        if numero:
            qs = qs.filter(orca_nume=numero)

        if status not in (None, ""):
            qs = qs.filter(orca_stat=status)

        if data_inicio:
            try:
                di = datetime.strptime(data_inicio, '%Y-%m-%d').date()
                qs = qs.filter(orca_data__gte=di)
            except ValueError:
                pass

        if data_fim:
            try:
                df = datetime.strptime(data_fim, '%Y-%m-%d').date()
                qs = qs.filter(orca_data__lte=df)
            except ValueError:
                pass

        if cliente_nome:
            from Entidades.models import Entidades
            clientes_ids = Entidades.objects.using(self.banco).filter(
                enti_nome__icontains=cliente_nome
            ).values_list("enti_clie", flat=True)
            qs = qs.filter(orca_clie__in=list(clientes_ids))

        if vendedor_nome:
            from Entidades.models import Entidades
            from django.db.models import Q
            vendedor_nomes = self.request.GET.getlist("vendedor_nome")
            q_objects = Q()
            for nome in vendedor_nomes:
                if nome:
                    q_objects |= Q(enti_nome__icontains=nome)
            vendedores_ids = Entidades.objects.using(self.banco).filter(q_objects).values_list("enti_clie", flat=True)
            qs = qs.filter(orca_vend__in=list(vendedores_ids))

        if empresa_id:
            try:
                qs = qs.filter(orca_empr=int(empresa_id))
            except ValueError:
                pass

        if filial_id:
            try:
                qs = qs.filter(orca_fili=int(filial_id))
            except ValueError:
                pass

        context["slug"] = self.kwargs["slug"]

        # Get list of vendedores for the dropdown
        from Entidades.models import Entidades
        vendedores = Entidades.objects.using(self.banco).filter(
            enti_tipo_enti='VE'
        ).values('enti_clie', 'enti_nome').order_by('enti_nome')
        context["vendedores_list"] = list(vendedores)

        # Get empresas and filiais for filters
        try:
            from Licencas.models import Empresas, Filiais
            context["empresas_list"] = list(
                Empresas.objects.using(self.banco).only("empr_codi", "empr_nome").order_by("empr_nome")
            )
            filiais_qs = Filiais.objects.using(self.banco).only("empr_empr", "empr_codi", "empr_nome").order_by("empr_nome")
            if empresa_id:
                try:
                    filiais_qs = filiais_qs.filter(empr_empr=int(empresa_id))
                except ValueError:
                    pass
            context["filiais_list"] = list(filiais_qs)
        except Exception:
            context["empresas_list"] = []
            context["filiais_list"] = []

        # Calculate metrics based on filtered queryset
        context["metricas"] = {
            "total_orcamentos": qs.count(),
            "total_valor": qs.aggregate(total=Sum("orca_tota")).get("total") or 0,
            "total_exportados": qs.filter(orca_stat=2).count(),
            "total_cancelados": qs.filter(orca_stat=3).count(),
        }

        # Calculate totals by empresa/filial
        empresa_totals = qs.values('orca_empr').annotate(
            total=Sum('orca_tota'),
            count=Count('orca_nume')
        ).order_by('-total')
        
        # Get empresa names
        empresa_dict = {}
        try:
            from Licencas.models import Empresas
            for item in empresa_totals:
                if item['orca_empr']:
                    try:
                        emp = Empresas.objects.using(self.banco).filter(empr_codi=item['orca_empr']).first()
                        if emp:
                            empresa_dict[item['orca_empr']] = emp.empr_nome
                    except:
                        pass
        except:
            pass
        
        context["empresa_totals"] = [
            {
                'empresa_id': item['orca_empr'],
                'empresa_nome': empresa_dict.get(item['orca_empr'], f"Empresa {item['orca_empr']}"),
                'total': item['total'] or 0,
                'count': item['count']
            }
            for item in empresa_totals if item['orca_empr']
        ]

        # Set default dates to current month if not provided
        today = date.today()
        first_day_of_month = date(today.year, today.month, 1)

        # Handle multiple vendedor_nome values (always return a list)
        vendedor_nome_filter = self.request.GET.getlist("vendedor_nome")

        context["filtros"] = {
            "orca_nume": self.request.GET.get("orca_nume", ""),
            "cliente_nome": self.request.GET.get("cliente_nome", ""),
            "vendedor_nome": vendedor_nome_filter,
            "orca_stat": self.request.GET.get("orca_stat", ""),
            "data_inicio": self.request.GET.get("data_inicio", first_day_of_month.strftime('%Y-%m-%d')),
            "data_fim": self.request.GET.get("data_fim", today.strftime('%Y-%m-%d')),
            "empresa_id": empresa_id or "",
            "filial_id": filial_id or "",
        }

        return context


class ExportarOrcamentosView(View):
    def get(self, request, *args, **kwargs):
        slug = kwargs.get("slug")
        db_alias = get_db_from_slug(slug)
        data_min = date(2020, 1, 1)

        # Build queryset with filters
        qs = (
            Orcamentopisos.objects
            .using(db_alias)
            .filter(orca_data__gte=data_min)
            .only(
                "orca_empr",
                "orca_fili",
                "orca_nume",
                "orca_clie",
                "orca_vend",
                "orca_data",
                "orca_tota",
                "orca_stat",
            )
            .order_by("-orca_nume")
        )

        # Apply vendor filter
        mix = VendedorEntidadeMixin()
        mix.request = request
        qs = mix.filter_por_vendedor(qs, "orca_vend")

        # Apply filters from request
        numero = request.GET.get("orca_nume")
        cliente_nome = request.GET.get("cliente_nome")
        vendedor_nome = request.GET.get("vendedor_nome")
        status = request.GET.get("orca_stat")
        data_inicio = request.GET.get("data_inicio")
        data_fim = request.GET.get("data_fim")

        if numero:
            qs = qs.filter(orca_nume=numero)

        if status not in (None, ""):
            qs = qs.filter(orca_stat=status)

        if data_inicio:
            try:
                di = datetime.strptime(data_inicio, '%Y-%m-%d').date()
                qs = qs.filter(orca_data__gte=di)
            except ValueError:
                pass

        if data_fim:
            try:
                df = datetime.strptime(data_fim, '%Y-%m-%d').date()
                qs = qs.filter(orca_data__lte=df)
            except ValueError:
                pass

        if cliente_nome:
            clientes_ids = Entidades.objects.using(db_alias).filter(
                enti_nome__icontains=cliente_nome
            ).values_list("enti_clie", flat=True)
            qs = qs.filter(orca_clie__in=list(clientes_ids))

        if vendedor_nome:
            vendedor_nomes = request.GET.getlist("vendedor_nome")
            from django.db.models import Q
            q_objects = Q()
            for nome in vendedor_nomes:
                if nome:
                    q_objects |= Q(enti_nome__icontains=nome)
            vendedores_ids = Entidades.objects.using(db_alias).filter(q_objects).values_list("enti_clie", flat=True)
            qs = qs.filter(orca_vend__in=list(vendedores_ids))

        rows = list(qs)

        # Get entity names
        entidades_ids = set()
        for o in rows:
            if o.orca_clie:
                entidades_ids.add(o.orca_clie)
            if o.orca_vend:
                entidades_ids.add(o.orca_vend)

        entidades = Entidades.objects.using(db_alias).filter(
            enti_clie__in=entidades_ids
        )
        nomes = {e.enti_clie: e.enti_nome for e in entidades}

        # Get status map
        if rows:
            empresa = rows[0].orca_empr
            filial = rows[0].orca_fili
            status_map = StatusPisosService.mapa_status(
                banco=db_alias,
                empresa=empresa,
                filial=filial,
                tipo=StatusPisos.TIPO_ORCAMENTO,
            )
        else:
            status_map = {}

        # Calculate totals
        total_orcamentos = len(rows)
        total_valor = sum((o.orca_tota or 0) for o in rows)
        total_exportados = sum((o.orca_tota or 0) for o in rows if o.orca_stat == 2)  # Status 2 = exportado

        # Create Excel response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=orcamentos.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = 'Orçamentos'

        headers = [
            'Número',
            'Data',
            'Cliente',
            'Vendedor',
            'Status',
            'Total'
        ]
        ws.append(headers)

        # Format header
        header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Add data rows
        for o in rows:
            status_obj = status_map.get(o.orca_stat)
            status_desc = status_obj.stat_desc if status_obj else "Sem status"

            ws.append([
                o.orca_nume or "",
                o.orca_data.strftime("%d/%m/%Y") if o.orca_data else "",
                nomes.get(o.orca_clie, ""),
                nomes.get(o.orca_vend, ""),
                status_desc,
                float(o.orca_tota or 0),
            ])

        # Add totals row
        ws.append([])
        ws.append([
            "TOTAIS",
            "",
            "",
            "",
            f"Orçamentos: {total_orcamentos}",
            float(total_valor),
        ])
        ws.append([
            "",
            "",
            "",
            "",
            f"Exportados: {sum(1 for o in rows if o.orca_stat == 2)}",
            float(total_exportados),
        ])

        # Format totals row
        total_fill = PatternFill(fill_type='solid', fgColor='4472C4')
        total_font = Font(color='FFFFFF', bold=True)
        for row_num in [ws.max_row, ws.max_row - 1]:
            for cell in ws[row_num]:
                cell.fill = total_fill
                cell.font = total_font

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(response)
        return response