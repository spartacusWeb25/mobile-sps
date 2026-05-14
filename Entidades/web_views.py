import logging
from django.shortcuts import render, redirect
from django.contrib import messages

logger = logging.getLogger(__name__)
from django.utils import timezone
from django.http import Http404, HttpResponse, JsonResponse
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, TemplateView
from Agricola.service.cadastros_service import CadastrosDomainService
from .services.entidades_trasportadores import EntidadeTransportadoraServico
from .services.entidades_motoristas import EntidadeMotoristaServico
from urllib.parse import quote_plus
from django.db.models import BigIntegerField, Case, When, Value, CharField, IntegerField, OuterRef, Subquery
from django.db.models.functions import Cast
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Entidades
from .forms import EntidadesForm
from core.utils import get_licenca_db_config


class DBAndSlugMixin:
    slug_url_kwarg = 'slug'

    def dispatch(self, request, *args, **kwargs):
        db_alias = get_licenca_db_config(request)
        setattr(request, 'db_alias', db_alias)
        # Capturar empresa/filial priorizando sessão; fallback para headers e querystring
        self.empresa_id = (
            request.session.get('empresa_id')
            or request.headers.get('X-Empresa')
            or request.GET.get('enti_empr')
        )
        self.filial_id = (
            request.session.get('filial_id')
            or request.headers.get('X-Filial')
            or request.GET.get('enti_fili')
        )
        self.slug = kwargs.get(self.slug_url_kwarg)
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['slug'] = getattr(self, 'slug', None)
        context['current_year'] = timezone.now().year
        return context


class EntidadeListView(DBAndSlugMixin, ListView):
    template_name = 'Entidades/entidades.html'
    context_object_name = 'entidades'
    paginate_by = 18

    def get_queryset(self):
        request = self.request
        db_alias = getattr(request, 'db_alias', None)
        qs = Entidades.objects.using(db_alias).all()
        # Filtrar por empresa quando disponível para evitar duplicidades
        if self.empresa_id:
            qs = qs.filter(enti_empr=int(self.empresa_id))
        qs = qs.order_by('enti_empr', 'enti_nome')
        nome = request.GET.get('enti_nome', '')
        id_cliente = request.GET.get('enti_clie', '')
        tipo = request.GET.get('enti_tipo_enti', '')
        classificacao = request.GET.get('enti_espe_enti', '')
        vendedor_responsavel = request.GET.get('enti_vend', '')
        situacao = request.GET.get('enti_situ', '')
        
        if tipo:
            qs = qs.filter(enti_tipo_enti__icontains=tipo)
        if classificacao:
            qs = qs.filter(enti_espe_enti=classificacao)
        if situacao:
            qs = qs.filter(enti_situ__icontains=situacao)
        if nome:
            qs = qs.filter(enti_nome__icontains=nome)
        if id_cliente:
            try:
                qs = qs.filter(enti_clie=int(id_cliente))
            except (ValueError, TypeError):
                pass
        if vendedor_responsavel:
            try:
                qs = qs.filter(enti_vend=int(vendedor_responsavel))
            except (ValueError, TypeError):
                pass

        vendedores_sub = Entidades.objects.using(db_alias).filter(
            enti_clie=Cast(OuterRef("enti_vend"), BigIntegerField())
        )
        if self.empresa_id:
            vendedores_sub = vendedores_sub.filter(enti_empr=int(self.empresa_id))

        qs = qs.annotate(
            vendedor_responsavel_nome=Subquery(
                vendedores_sub.values("enti_nome")[:1]
            )
        )
        return qs

    def get_context_data(self, **kwargs):
        qs = self.get_queryset()
        context = super().get_context_data(**kwargs)
        request = self.request
        nome = request.GET.get('enti_nome', '')
        id_cliente = request.GET.get('enti_clie', '')
        tipo = request.GET.get('enti_tipo_enti', '')
        classificacao = request.GET.get('enti_espe_enti', '')
        situacao = request.GET.get('enti_situ', '')
        vendedor_responsavel = request.GET.get('enti_vend', '')
        db_alias = getattr(request, 'db_alias', None)
        vendedor_responsavel_nome = CadastrosDomainService.vendedor_nome_por_enti_clie(vendedor_responsavel, db_alias)
        
        total_entidades = qs.count()
        total_de_clientes = qs.filter(enti_tipo_enti='CL').count()
        total_de_fornecedores = qs.filter(enti_tipo_enti='FO').count()
        total_de_motoristas = qs.filter(enti_tipo_enti='FU').count()
        total_de_ambos = qs.filter(enti_tipo_enti__in=['AM']).count()

        context['nome'] = nome
        context['id_cliente'] = id_cliente
        context['tipo_selecionado'] = tipo
        context['classificacao_selecionada'] = classificacao
        context['situacao_selecionada'] = situacao
        context['total_entidades'] = total_entidades
        context['total_de_clientes'] = total_de_clientes
        context['total_de_fornecedores'] = total_de_fornecedores
        context['total_de_motoristas'] = total_de_motoristas
        context['total_de_ambos'] = total_de_ambos
        
        # Opções para os filtros
        context['tipos_entidade'] = Entidades.TIPO_ENTIDADES
        context['classificacoes_tributarias'] = Entidades.CLASSIFICACAO_TRIBUTACAO
        context['situacoes_entidade'] = Entidades._meta.get_field('enti_situ').choices

        # Preservar filtros na paginação
        extra_parts = []
        if nome:
            extra_parts.append('&enti_nome=' + quote_plus(nome))
        if id_cliente:
            extra_parts.append('&enti_clie=' + quote_plus(id_cliente))
        if tipo:
            extra_parts.append('&enti_tipo_enti=' + quote_plus(tipo))
        if classificacao:
            extra_parts.append('&enti_espe_enti=' + quote_plus(classificacao))
        if situacao:
            extra_parts.append('&enti_situ=' + quote_plus(situacao))
        if vendedor_responsavel:
            extra_parts.append('&enti_vend=' + quote_plus(vendedor_responsavel))
            
        context['extra_query'] = ''.join(extra_parts)
        context['vendedor_responsavel_nome'] = vendedor_responsavel_nome
        context['vendedor_responsavel_selecionado'] = int(vendedor_responsavel) if str(vendedor_responsavel or '').strip().isdigit() else None
        try:
            empresa_id = int(self.empresa_id) if self.empresa_id not in [None, ""] else None
        except Exception:
            empresa_id = None
        vendedores_qs = Entidades.objects.using(db_alias).filter(
            enti_tipo_enti__in=["VE", "AM", "FU"],
            enti_situ="1",
        )
        if empresa_id is not None:
            vendedores_qs = vendedores_qs.filter(enti_empr=empresa_id)
        vendedores_qs = vendedores_qs.only("enti_clie", "enti_nome").order_by("enti_nome")[:500]
        context['vendedores_entidade'] = [(v.enti_clie, f"{v.enti_clie} - {v.enti_nome}") for v in vendedores_qs]
        return context


def autocomplete_vendedores(request, slug=None):
    banco = get_licenca_db_config(request) or "default"
    empresa_id = request.session.get("empresa_id") or request.headers.get("X-Empresa")
    term = (request.GET.get("term") or request.GET.get("q") or "").strip()

    try:
        empresa_id = int(empresa_id)
    except (TypeError, ValueError):
        return JsonResponse({"results": []})

    qs = Entidades.objects.using(banco).filter(
        enti_empr=empresa_id,
        enti_tipo_enti__in=["VE", "AM", "FU"],
        enti_situ="1",
    )
    if term:
        if term.isdigit():
            qs = qs.filter(enti_clie__icontains=term)
        else:
            qs = qs.filter(enti_nome__icontains=term)
    qs = qs.only("enti_clie", "enti_nome").order_by("enti_nome")[:20]

    data = [{"id": str(v.enti_clie), "text": f"{v.enti_clie} - {v.enti_nome}"} for v in qs]
    return JsonResponse({"results": data})


class EntidadeCreateView(DBAndSlugMixin, CreateView):
    template_name = 'Entidades/entidade_form.html'
    form_class = EntidadesForm

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    

    def form_valid(self, form):
        logger.info("EntidadeCreateView.form_valid chamado.")
        db_alias = getattr(self.request, 'db_alias', 'default')
        try:
            self.object = self.execute_create(form, db_alias)
            logger.info(f"Entidade criada com sucesso: {self.object}")
        except Exception as e:
            logger.error(f"Erro no execute_create: {e}", exc_info=True)
            form.add_error(None, f"Erro ao cadastrar entidade: {e}")
            return self.form_invalid(form)

        messages.success(self.request, 'Entidade cadastrada com sucesso.')
        return redirect(self.get_success_url())

    def get_success_url(self):
        return reverse_lazy('entidades_web', kwargs={'slug': self.slug})

    def execute_create(self, form, db_name):
        logger.info("execute_create iniciado.")
        data = form.cleaned_data.copy()        
        
        # Tenta obter empresa/filial de várias fontes
        empresa = self.empresa_id
        filial = self.filial_id

        # Fallback para request.user (comum em outros módulos)
        if not empresa and self.request.user.is_authenticated:
            empresa_user = getattr(self.request.user, 'empresa', None)
            if hasattr(empresa_user, 'id'):
                empresa = empresa_user.id
            elif empresa_user:
                empresa = empresa_user
            
            if empresa:
                logger.info(f"Empresa recuperada de request.user: {empresa}")

        # Fallback direto para session se ainda não encontrou
        if not empresa:
            empresa = self.request.session.get('empresa_id')
            if empresa:
                logger.info(f"Empresa recuperada diretamente da session: {empresa}")

        # Fallback para filial
        if not filial and self.request.user.is_authenticated:
            filial_user = getattr(self.request.user, 'filial', None)
            if hasattr(filial_user, 'id'):
                filial = filial_user.id
            elif filial_user:
                filial = filial_user
            
        if not filial:
             filial = self.request.session.get('filial_id', 1)

        # Garante que são inteiros
        try:
            if empresa: empresa = int(empresa)
            if filial: filial = int(filial)
        except (ValueError, TypeError):
             logger.warning(f"Falha ao converter empresa/filial para int: {empresa}/{filial}")

        logger.info(f"Empresa final: {empresa}, Filial final: {filial}")
        
        # Debug da sessão se falhar
        if not empresa:
            logger.error(f"Sessão keys: {list(self.request.session.keys())}")
            try:
                 logger.error(f"User dir: {dir(self.request.user)}")
                 logger.error(f"User is_authenticated: {self.request.user.is_authenticated}")
            except:
                 pass
            
        if not empresa and 'enti_empr' in data:
            empresa = data['enti_empr']
            logger.info(f"Empresa recuperada do form data: {empresa}")

        # ÚLTIMA TENTATIVA: Header X-Empresa vindo do frontend (HTMX/Ajax)
        if not empresa:
            empresa = self.request.headers.get('X-Empresa')
            if empresa:
                logger.info(f"Empresa recuperada do Header X-Empresa: {empresa}")
            
        if not empresa:
            # Tenta pegar da URL se for GET ou query params no POST
            empresa = self.request.GET.get('enti_empr')
            if empresa:
                 logger.info(f"Empresa recuperada do request.GET: {empresa}")

        if not empresa:
            logger.error("Empresa não identificada.")
            raise ValueError("Empresa não identificada para o cadastro.")

        data['enti_empr'] = empresa
        
        if form.cleaned_data.get('is_transportadora'):
             return EntidadeTransportadoraServico.cadastrar_transportadora(
                data=data,
                empresa_id=empresa,
                filial_id=filial,
                banco=db_name
            )
        
        if form.cleaned_data.get('is_motorista'):
             return EntidadeMotoristaServico.cadastrar_motorista(
                data=data,
                empresa_id=empresa,
                filial_id=filial,
                banco=db_name
            )

        return CadastrosDomainService.cadastrar_entidade(
            empresa=empresa,
            filial=filial,
            dados=data,
            using=db_name
        )

class EntidadeUpdateView(DBAndSlugMixin, UpdateView):
    template_name = 'Entidades/entidade_form.html'
    form_class = EntidadesForm
    model = Entidades
    pk_url_kwarg = 'enti_clie'

    def get_queryset(self):
        db_alias = getattr(self.request, 'db_alias', None)
        qs = Entidades.objects.using(db_alias).all()
        if self.empresa_id:
            qs = qs.filter(enti_empr=int(self.empresa_id))
        return qs

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def get_success_url(self):
        return reverse_lazy('entidades_web', kwargs={'slug': self.slug})


class EntidadeDeleteView(DBAndSlugMixin, DeleteView):
    template_name = 'Entidades/entidade_confirm_delete.html'
    model = Entidades
    pk_url_kwarg = 'enti_clie'

    def get_queryset(self):
        db_alias = getattr(self.request, 'db_alias', None)
        qs = Entidades.objects.using(db_alias).all()
        if self.empresa_id:
            qs = qs.filter(enti_empr=int(self.empresa_id))
        return qs

    def delete(self, request, *args, **kwargs):
        try:
            return super().delete(request, *args, **kwargs)
        except Exception as e:
            messages.error(request, f'Erro ao excluir: {e}')
            return redirect('entidades_web', slug=self.slug)

    def get_success_url(self):
        messages.success(self.request, 'Entidade excluída com sucesso.')
        return reverse_lazy('entidades_web', kwargs={'slug': self.slug})


class ExportarEntidadesView(DBAndSlugMixin, View):
    def get(self, request, *args, **kwargs):
        db_alias = getattr(request, 'db_alias', None)
        nome = request.GET.get('enti_nome', '')
        id_cliente = request.GET.get('enti_clie', '')
        queryset = (
            Entidades.objects.using(db_alias)
            .all()
            .annotate(
                enti_clie_int=Cast('enti_clie', IntegerField()),
                tipo=Case(
                    When(enti_tipo_enti='CL', then=Value('Cliente')),
                    When(enti_tipo_enti='FO', then=Value('Fornecedor')),
                    When(enti_tipo_enti='AM', then=Value('Ambos')),
                    When(enti_tipo_enti='FU', then=Value('Funcionário')),
                    default=Value('Desconhecido'),
                    output_field=CharField(),
                ),
                tipo_formatado=Case(
                    When(enti_tien='T', then=Value('Transportadora')),
                    When(enti_tien='M', then=Value('Motorista')),
                    default=Value('Entidade'),
                    output_field=CharField(),
                ),
                situacao_formatada=Case(
                    When(enti_situ='1', then=Value('Ativo')),
                    When(enti_situ='0', then=Value('Inativo')),
                    default=Value('Desconhecido'),
                    output_field=CharField(),
                ),
            )
            .order_by('enti_clie_int')
        )

        if nome:
            queryset = queryset.filter(enti_nome__icontains=nome)

        if id_cliente:
            try:
                queryset = queryset.filter(enti_clie_int=int(id_cliente))
            except (ValueError, TypeError):
                pass

        wb = Workbook()
        ws = wb.active
        ws.title = 'Entidades'

        headers = [
            'ID', 'Nome', 'Classificação', 'CPF', 'CNPJ',
            'Cidade', 'Estado', 'Telefone', 'Celular',
            'Email', 'Situação', 'Tipo'
        ]
        ws.append(headers)

        # Cabeçalho formatado
        header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center')

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for e in queryset:
            ws.append([
                e.enti_clie or '',
                e.enti_nome or '',
                e.tipo or '',
                e.enti_cpf or '',
                e.enti_cnpj or '',
                e.enti_cida or '',
                e.enti_esta or '',
                e.enti_fone or '',
                e.enti_celu or '',
                e.enti_emai or '',
                e.situacao_formatada or '',
                e.tipo_formatado or '',
            ])

        # Ajuste simples de largura
        larguras = {
            'A': 12, 'B': 35, 'C': 18, 'D': 18, 'E': 20,
            'F': 20, 'G': 10, 'H': 18, 'I': 18, 'J': 30,
            'K': 15, 'L': 18,
        }
        for col, largura in larguras.items():
            ws.column_dimensions[col].width = largura

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="entidades.xlsx"'

        wb.save(response)
        return response



class RelatorioEntidadesView(TemplateView):
    template_name = "Entidades/relatorio_entidades.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        try:
            from Licencas.models import Empresas
            ctx["empresas"] = list(Empresas.objects.all().values("empr_codi", "empr_nome"))
        except Exception:
            ctx["empresas"] = []
        return ctx
