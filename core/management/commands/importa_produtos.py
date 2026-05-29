from django.core.management.base import BaseCommand, CommandError
from django.db import connections
from django.db.utils import OperationalError

from core.licencas_loader import carregar_licencas_dict
from openpyxl import load_workbook
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor
import pandas as pd
from PIL import Image, ImageFilter
import io
import time


# ========= CONFIGURAÇÕES =========
EXCEL_PATH      = "KOHLER.xlsx"
COL_CODIGO      = 4        # Coluna D → CÓDIGO FG
COL_DESC        = 7        # Coluna G → DESCRIÇÃO
BUCKET_NAME     = "produtos"
NAMESPACE       = "grsxg5eatn7l"
REGION          = "sa-saopaulo-1"
OUTPUT_CSV      = "produtos_com_url.csv"

SUFIXOS_COR = {'BN','CP','RGD','TT','BL','VS','NA','G','A','B','C4','K'}

TARGET_MIN_SIZE = 800
UPSCALE_FACTOR  = 2
# =================================


def montar_db_config(lic):
    config = {
        "ENGINE": "django.db.backends.postgresql",
        "NAME": lic["db_name"],
        "USER": lic["db_user"],
        "PASSWORD": lic["db_password"],
        "HOST": lic["db_host"],
        "PORT": lic["db_port"],
        "CONN_MAX_AGE": 60,
    }
    return config


def get_prefix(codigo: str) -> str:
    parts = codigo.strip().split('-')
    if len(parts) > 1 and parts[-1].upper() in SUFIXOS_COR:
        return '-'.join(parts[:-1])
    return '-'.join(parts)


class Command(BaseCommand):
    help = "Importa produtos do Excel KOHLER e envia imagens para o Oracle Object Storage"

    def add_arguments(self, parser):
        parser.add_argument("--preview", action="store_true",
                            help="Mostra o que será feito sem enviar nada ao bucket")
        parser.add_argument("--retry", type=int, default=3,
                            help="Número de tentativas no upload (default: 3)")
        parser.add_argument(
            "--slug",
            type=str,
            help="Slug do tenant específico. Se omitido, roda em todos os tenants.",
        )
        parser.add_argument(
            "--tenant",
            type=str,
            help="Alias de --slug (compatibilidade).",
        )

    def process_image(self, raw_bytes: bytes) -> bytes | None:
        try:
            image = Image.open(io.BytesIO(raw_bytes))
            if image.mode != "RGBA":
                image = image.convert("RGBA")
            w, h = image.size
            if w < TARGET_MIN_SIZE or h < TARGET_MIN_SIZE:
                new_w, new_h = int(w * UPSCALE_FACTOR), int(h * UPSCALE_FACTOR)
                image = image.resize((new_w, new_h), Image.Resampling.LANCZOS)
                self.stdout.write(f"      ↑ upscale {w}×{h} → {new_w}×{new_h}")
            else:
                self.stdout.write(f"      ✓ {w}×{h}px")
            image = image.filter(ImageFilter.UnsharpMask(radius=1.5, percent=130, threshold=3))
            output = io.BytesIO()
            image.save(output, format="PNG", optimize=True)
            return output.getvalue()
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"      Erro ao processar imagem: {e}"))
            return None

    def upload(self, object_storage, obj_path: str, data: bytes, max_retries: int) -> bool:
        for attempt in range(1, max_retries + 1):
            try:
                object_storage.put_object(
                    namespace_name=NAMESPACE, bucket_name=BUCKET_NAME,
                    object_name=obj_path, put_object_body=data, content_type="image/png",
                )
                return True
            except Exception as e:
                self.stdout.write(self.style.WARNING(f"      tentativa {attempt}/{max_retries}: {e}"))
                if attempt < max_retries:
                    time.sleep(2 ** attempt)
        return False

    def importar_produtos(self, slug: str, preview: bool, max_retry: int):
        """Executa a importação de produtos para um tenant específico."""
        self.stdout.write(self.style.WARNING("═" * 65))
        self.stdout.write(self.style.WARNING(f"  Tenant: {slug}"))
        self.stdout.write(self.style.WARNING(
            "  PREVIEW — nenhum upload será feito" if preview
            else "  EXECUÇÃO REAL — imagens serão enviadas ao OCI"
        ))
        self.stdout.write(self.style.WARNING("═" * 65))

        object_storage = None
        if not preview:
            import oci
            config = oci.config.from_file()
            object_storage = oci.object_storage.ObjectStorageClient(config)

        try:
            wb = load_workbook(EXCEL_PATH)
            ws = wb.active
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR(f"Arquivo não encontrado: {EXCEL_PATH}"))
            return
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Erro ao abrir Excel: {e}"))
            return

        self.stdout.write(f"\nImagens no Excel: {len(ws._images)}")
        self.stdout.write(f"Linhas de dados:  {ws.max_row - 1}\n")

        # ------------------------------------------------------------------
        # PASSO 1 — Pré-carrega TODOS os bytes de imagem em memória.
        #
        # CRÍTICO: img._data() fecha o file handle após a primeira leitura.
        # Se a mesma imagem for referenciada por herança em múltiplos produtos,
        # a segunda chamada lança "I/O operation on closed file" e o produto
        # fica sem URL. Carregar tudo em bytes[] logo aqui elimina esse problema.
        # ------------------------------------------------------------------
        linha_para_bytes: dict[int, bytes] = {}
        for img in ws._images:
            anchor = img.anchor
            try:
                if isinstance(anchor, TwoCellAnchor):
                    r_start = anchor._from.row + 1
                    r_end   = anchor.to.row + 1
                elif isinstance(anchor, OneCellAnchor):
                    r_start = r_end = anchor._from.row + 1
                elif hasattr(anchor, "_from"):
                    r_start = r_end = anchor._from.row + 1
                else:
                    continue
            except AttributeError:
                continue
            try:
                raw = img._data()   # única leitura — carrega agora, guarda para sempre
                if not raw:
                    continue
            except Exception:
                continue
            for r in range(r_start, r_end + 1):
                linha_para_bytes[r] = raw

        self.stdout.write(f"Linhas com imagem mapeadas: {len(linha_para_bytes)}")

        # ------------------------------------------------------------------
        # PASSO 2 — Mapeia prefixo → bytes da imagem PRÓPRIA (primeira direta).
        #
        # Só registra quando a linha tem imagem ancorada, para não herdar
        # erroneamente a imagem do produto anterior.
        # ------------------------------------------------------------------
        prefixo_bytes_proprios: dict[str, bytes] = {}
        for r in range(2, ws.max_row + 1):
            if r not in linha_para_bytes:
                continue
            codigo = ws.cell(row=r, column=COL_CODIGO).value
            if not codigo:
                continue
            prefixo = get_prefix(str(codigo).strip())
            if prefixo not in prefixo_bytes_proprios:
                prefixo_bytes_proprios[prefixo] = linha_para_bytes[r]

        self.stdout.write(f"Prefixos com imagem própria: {len(prefixo_bytes_proprios)}\n")

        # ------------------------------------------------------------------
        # PASSO 3 — Percorre todos os produtos e resolve bytes finais.
        #
        # Prioridade:
        #   1. Bytes próprios do prefixo (prefixo_bytes_proprios)
        #   2. Herança: last_bytes — últimos bytes vistos de cima p/ baixo
        #
        # Como tudo está em memória (bytes), nunca há I/O closed file.
        # ------------------------------------------------------------------
        produtos_csv      = []
        prefixos_enviados: dict[str, str] = {}   # prefixo → URL
        uploads_falhos    = []
        last_bytes: bytes | None = None

        for r in range(2, ws.max_row + 1):
            codigo = ws.cell(row=r, column=COL_CODIGO).value
            if not codigo:
                continue

            codigo    = str(codigo).strip()
            descricao = str(ws.cell(row=r, column=COL_DESC).value or "").strip()
            prefixo   = get_prefix(codigo)
            nome_arq  = f"{prefixo}.png"
            obj_path  = f"produtos/{nome_arq}"
            url       = (
                f"https://objectstorage.{REGION}.oraclecloud.com"
                f"/n/{NAMESPACE}/b/{BUCKET_NAME}/o/{obj_path}"
            )

            # Atualiza last_bytes se esta linha tem imagem própria
            if r in linha_para_bytes:
                last_bytes = linha_para_bytes[r]

            # Resolve bytes: próprio do prefixo > herança de last_bytes
            raw   = prefixo_bytes_proprios.get(prefixo, last_bytes)
            fonte = "própria" if prefixo in prefixo_bytes_proprios else "herdada"

            ja_enviado = prefixo in prefixos_enviados

            self.stdout.write(
                f"Linha {r:4d} | {codigo:30s} | {nome_arq:40s} | {fonte}"
                + (" [já enviado]" if ja_enviado else "")
            )

            if not ja_enviado:
                if not preview:
                    if raw is None:
                        self.stdout.write(self.style.ERROR("      sem imagem disponível"))
                        produtos_csv.append({"codigo": codigo, "descricao": descricao, "prod_url": ""})
                        continue
                    processed = self.process_image(raw)
                    if not processed:
                        produtos_csv.append({"codigo": codigo, "descricao": descricao, "prod_url": ""})
                        continue
                    ok = self.upload(object_storage, obj_path, processed, max_retry)
                    if ok:
                        self.stdout.write(self.style.SUCCESS(f"      ✔ upload ok → {nome_arq}"))
                        prefixos_enviados[prefixo] = url
                    else:
                        self.stdout.write(self.style.ERROR("      ✘ upload falhou"))
                        uploads_falhos.append(prefixo)
                        url = ""
                else:
                    # preview: registra a URL sem fazer upload
                    if raw is None:
                        self.stdout.write(self.style.WARNING("      [sem imagem]"))
                        url = ""
                    prefixos_enviados[prefixo] = url

            produtos_csv.append({
                "codigo":    codigo,
                "descricao": descricao,
                "prod_url":  prefixos_enviados.get(prefixo, ""),
            })

        # ------------------------------------------------------------------
        # RESUMO
        # ------------------------------------------------------------------
        self.stdout.write("\n" + self.style.WARNING("═" * 65))
        self.stdout.write(self.style.WARNING("  RESUMO"))
        self.stdout.write(self.style.WARNING("═" * 65))
        self.stdout.write(f"  Produtos no CSV:            {len(produtos_csv)}")
        self.stdout.write(f"  Arquivos únicos no bucket:  {len(prefixos_enviados)}")
        sem_url = sum(1 for p in produtos_csv if not p["prod_url"])
        if sem_url:
            self.stdout.write(self.style.ERROR(f"  Produtos sem URL:           {sem_url}"))
            for p in produtos_csv:
                if not p["prod_url"]:
                    self.stdout.write(self.style.ERROR(f"    → {p['codigo']}"))
        else:
            self.stdout.write(self.style.SUCCESS("  Todos os 651 produtos com URL ✓"))
        if uploads_falhos:
            self.stdout.write(self.style.ERROR(
                f"  Falhas ({len(uploads_falhos)}): {', '.join(uploads_falhos)}"
            ))

        output_csv = f"produtos_com_url_{slug}.csv"
        df = pd.DataFrame(produtos_csv)
        df.to_csv(output_csv, index=False, encoding="utf-8-sig")
        self.stdout.write(self.style.SUCCESS(f"\n  CSV gerado: {output_csv}"))

        if preview:
            self.stdout.write(self.style.WARNING("\n  Nenhum arquivo foi enviado (modo preview)."))

    def handle(self, *args, **options):
        slug = options.get("slug")
        tenant = options.get("tenant")
        if slug and tenant and slug != tenant:
            raise CommandError("Use apenas um entre --slug e --tenant (ou informe o mesmo valor em ambos).")

        slug_alvo = slug or tenant
        preview = options["preview"]
        max_retry = options["retry"]

        licencas = carregar_licencas_dict()
        if not licencas:
            raise CommandError("Nenhuma licença encontrada")

        if slug_alvo:
            licencas = [l for l in licencas if l.get("slug") == slug_alvo]
            if not licencas:
                raise CommandError(f"Nenhuma licença encontrada para slug={slug_alvo}")

        for lic in licencas:
            alias = f"tenant_{lic['slug']}"
            connections.databases[alias] = montar_db_config(lic)

            try:
                with connections[alias].cursor() as cursor:
                    cursor.execute("SELECT 1")
            except OperationalError:
                self.stdout.write(self.style.ERROR(f"[{alias}] Banco de dados não encontrado ou inacessível. Pulando..."))
                continue

            self.stdout.write(self.style.WARNING(f"[{alias}] Iniciando importação de produtos..."))
            try:
                self.importar_produtos(lic['slug'], preview, max_retry)
                self.stdout.write(self.style.SUCCESS(f"[{alias}] Importação concluída com sucesso!"))
            except Exception as e:
                self.stdout.write(self.style.ERROR(f"[{alias}] Erro ao importar produtos: {e}"))