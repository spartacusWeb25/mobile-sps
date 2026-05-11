from django.views.generic import ListView, CreateView, UpdateView, DeleteView, View, TemplateView
from django.http import JsonResponse
from django.db.models import Q
from urllib.parse import quote_plus
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.http import HttpResponse, Http404
from django.contrib import messages
import logging
from django.db import transaction
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.db.models import DecimalField, IntegerField, OuterRef, Subquery
from django.db.models.functions import Cast, Coalesce
logger = logging.getLogger(__name__)

from core.utils import get_licenca_db_config
from core.middleware import get_licenca_slug
from django.db.models import Subquery, OuterRef, DecimalField, Value as V, IntegerField
from django.db.models.functions import Coalesce, Cast

from ...models import (
    Lote,
    Produtos,
    Tabelaprecos,
    TabelaprecosPromocional,
    TabelaprecosPromocionalhist,
    SaldoProduto,
    Tabelaprecoshist,
    GrupoProduto,
    SubgrupoProduto,
    FamiliaProduto,
    Marca,
    Ncm,
    UnidadeMedida,
)
from CFOP.models import CFOP as CFOPModel, NCM_CFOP_DIF, ProdutoFiscalPadrao
from django.utils import timezone
from ..prod_forms import (
    ProdutosForm,
    ProdutoFiscalPadraoForm,
    TabelaprecosFormSet,
    TabelaprecosPlainFormSet,
    TabelaprecosFormSetUpdate,
    TabelaprecosPromocionalFormSet,
    TabelaprecosPromocionalPlainFormSet,
    GrupoForm,
    SubgrupoForm,
    FamiliaForm,
    MarcaForm,
    UnidadeMedidaForm,
)
from ...servicos.preco_promocional import (
    atualizar_preco_com_historico as atualizar_preco_promocional_com_historico,
    criar_preco_com_historico as criar_preco_promocional_com_historico,
)
from CFOP.services.services import MotorFiscal
from CFOP.services.bases import FiscalContexto
import json
from decimal import Decimal
from Licencas.models import Filiais, Empresas
from CFOP.cst_utils import get_csts_por_regime




class DBAndSlugMixin:
    template_folder = 'templates_spsWeb/Produtos'

    def dispatch(self, request, *args, **kwargs):
        self.slug = kwargs.get('slug') or get_licenca_slug()
        self.db_alias = get_licenca_db_config(request)
        if not self.db_alias:
            raise Http404("Banco de dados da licença não encontrado")
        # Capturar empresa/filial priorizando sessão; fallback para headers e querystring
        self.empresa_id = (
            request.session.get('empresa_id')
            or request.headers.get('X-Empresa')
            or request.GET.get('prod_empr')
        )
        self.filial_id = (
            request.session.get('filial_id')
            or request.headers.get('X-Filial')
            or request.GET.get('prod_fili')
        )
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        # Direciona para a lista correspondente ao model da view
        try:
            mdl = getattr(self, 'model', None)
        except Exception:
            mdl = None
        from ...models import Produtos, GrupoProduto, SubgrupoProduto, FamiliaProduto, Marca
        mapping = {
            Produtos: 'produtos_web',
            GrupoProduto: 'grupos_web',
            SubgrupoProduto: 'subgrupos_web',
            FamiliaProduto: 'familias_web',
            Marca: 'marcas_web',
            UnidadeMedida: 'unidades_web',
        }
        name = mapping.get(mdl, 'produtos_web')
        return reverse_lazy(name, kwargs={'slug': self.slug or get_licenca_slug()})

class SimularImpostosView(DBAndSlugMixin, View):
    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)

            preco = Decimal(str(data.get('preco', '0')))
            ncm_codigo = data.get('ncm')
            
            # Construir objeto fiscal padrão mockado
            fiscal_mock = ProdutoFiscalPadrao()
            
            # Helper para converter string vazia ou None para None, e strings numéricas para Decimal
            def to_dec(val):
                if not val: return None
                return Decimal(str(val))
            
            fiscal_mock.cst_icms = data.get('cst_icms') or None
            fiscal_mock.aliq_icms = to_dec(data.get('aliq_icms'))
            
            fiscal_mock.cst_ipi = data.get('cst_ipi') or None
            fiscal_mock.aliq_ipi = to_dec(data.get('aliq_ipi'))
            
            fiscal_mock.cst_pis = data.get('cst_pis') or None
            fiscal_mock.aliq_pis = to_dec(data.get('aliq_pis'))
            
            fiscal_mock.cst_cofins = data.get('cst_cofins') or None
            fiscal_mock.aliq_cofins = to_dec(data.get('aliq_cofins'))
            
            fiscal_mock.cst_cbs = data.get('cst_cbs') or None
            fiscal_mock.aliq_cbs = to_dec(data.get('aliq_cbs'))
            
            fiscal_mock.cst_ibs = data.get('cst_ibs') or None
            fiscal_mock.aliq_ibs = to_dec(data.get('aliq_ibs'))
            
            # Construir contexto
            motor = MotorFiscal(banco=self.db_alias)
            
            # Resolver NCM
            ncm_obj = None
            if ncm_codigo:
                ncm_obj = Ncm.objects.using(self.db_alias).filter(ncm_codi=ncm_codigo).first()
            
            # Mock produto
            mock_produto = Produtos()
            mock_produto.fiscal = fiscal_mock # Attach fiscal override
            
            # Defaults para simulação
            # Tenta pegar da request ou usa defaults
            uf_origem = data.get('uf_origem') or 'SP' 
            uf_destino = data.get('uf_destino') or 'SP'
            tipo_oper = data.get('tipo_oper') or "VENDA" 
            
            empresa_id = int(self.empresa_id or 1)
            filial_id = int(self.filial_id or 1)
            
            try:
                filial = Filiais.objects.using(self.db_alias).filter(
                    empr_empr=empresa_id, 
                    empr_codi=filial_id
                ).first()
                regime = filial.empr_regi_trib if filial else '1'
            except Exception:
                regime = '1'
            
            ctx = FiscalContexto(
                empresa_id=empresa_id,
                filial_id=filial_id,
                banco=self.db_alias,
                regime=regime,
                uf_origem=uf_origem,
                uf_destino=uf_destino,
                produto=mock_produto,
                ncm=ncm_obj
            )
            
            resultado = motor.calcular_item(ctx, item=None, tipo_oper=tipo_oper, base_manual=preco)
            
            # Serialize Decimal
            def default_serializer(obj):
                if isinstance(obj, Decimal):
                    return str(obj)
                if hasattr(obj, 'cfop_codi'): # CFOP object
                     return obj.cfop_codi
                return str(obj)

            return JsonResponse(resultado, json_dumps_params={'default': default_serializer})
            
        except Exception as e:
            logger.error(f"Erro na simulação de impostos: {e}", exc_info=True)
            return JsonResponse({'error': str(e)}, status=400)

class ProdutoListView(DBAndSlugMixin, ListView):
    model = Produtos
    template_name = 'Produtos/produtos_lista.html'
    context_object_name = 'produtos'
    paginate_by = 20

    def get_queryset(self):
        qs = Produtos.objects.using(self.db_alias).all()
        # Filtrar por empresa se disponível para evitar duplicidades entre empresas
        if self.empresa_id:
            qs = qs.filter(prod_empr=str(self.empresa_id))
        prod_nome = (self.request.GET.get('prod_nome') or '').strip()
        prod_codi = (self.request.GET.get('prod_codi') or '').strip()
        if prod_nome:
            qs = qs.filter(prod_nome__icontains=prod_nome)
        if prod_codi:
            qs = qs.filter(prod_codi__icontains=prod_codi)
        # Anotar saldo de estoque via subquery (por empresa/filial quando disponíveis)
        saldo_qs = SaldoProduto.objects.using(self.db_alias).filter(
            produto_codigo=OuterRef('pk')
        )
        if self.empresa_id:
            saldo_qs = saldo_qs.filter(empresa=str(self.empresa_id))
        if self.filial_id:
            saldo_qs = saldo_qs.filter(filial=str(self.filial_id))
        saldo_sub = Subquery(saldo_qs.values('saldo_estoque')[:1], output_field=DecimalField())
        qs = qs.annotate(saldo_estoque=Coalesce(saldo_sub, V(0), output_field=DecimalField()))

        # Anotar preços principais (vista, prazo, custo) por empresa quando disponível
        preco_qs = Tabelaprecos.objects.using(self.db_alias).filter(
            tabe_prod=OuterRef('prod_codi')
        )
        if self.empresa_id:
            try:
                emp_int = int(self.empresa_id)
                preco_qs = preco_qs.filter(tabe_empr=emp_int)
            except Exception:
                pass
        if self.filial_id:
            try:
                fil_int = int(self.filial_id)
                preco_qs = preco_qs.filter(tabe_fili=fil_int)
            except Exception:
                pass
        preco_vista_sub = Subquery(preco_qs.values('tabe_avis')[:1], output_field=DecimalField())
        preco_prazo_sub = Subquery(preco_qs.values('tabe_praz')[:1], output_field=DecimalField())
        preco_custo_sub = Subquery(preco_qs.values('tabe_cuge')[:1], output_field=DecimalField())
        qs = qs.annotate(
            preco_vista=preco_vista_sub,
            preco_prazo=preco_prazo_sub,
            preco_custo=preco_custo_sub,
        )
        return qs.order_by('prod_empr', 'prod_codi')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        ctx['prod_nome'] = (self.request.GET.get('prod_nome') or '').strip()
        ctx['prod_codi'] = (self.request.GET.get('prod_codi') or '').strip()
        # Preservar filtros na paginação
        extra_parts = []
        if ctx['prod_nome']:
            extra_parts.append('&prod_nome=' + quote_plus(ctx['prod_nome']))
        if ctx['prod_codi']:
            extra_parts.append('&prod_codi=' + quote_plus(ctx['prod_codi']))
        ctx['extra_query'] = ''.join(extra_parts)
        # Popular pseudo-relacionamento de preços para o template existente (tabelaprecos_set.all)
        class _ManagerLike:
            def __init__(self, items):
                self._items = items
            def all(self):
                return self._items

        try:
            emp_int = int(self.empresa_id) if self.empresa_id else None
        except Exception:
            emp_int = None
        try:
            fil_int = int(self.filial_id) if self.filial_id else None
        except Exception:
            fil_int = None

        produtos = ctx.get('page_obj').object_list if ctx.get('page_obj') else ctx.get('produtos', [])
        for p in produtos:
            precos_qs = Tabelaprecos.objects.using(self.db_alias).filter(tabe_prod=p.prod_codi)
            if emp_int is not None:
                precos_qs = precos_qs.filter(tabe_empr=emp_int)
            if fil_int is not None:
                precos_qs = precos_qs.filter(tabe_fili=fil_int)
            p.tabelaprecos_set = _ManagerLike(list(precos_qs))
        return ctx


class ProdutoCreateView(DBAndSlugMixin, CreateView):
    model = Produtos
    form_class = ProdutosForm
    template_name = 'Produtos/produtos_form.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['database'] = self.db_alias
        return kwargs

    def _get_cst_choices(self):
        try:
            empresa_id = int(self.empresa_id or 1)
            filial_id = int(self.filial_id or 1)
            filial = Filiais.objects.using(self.db_alias).filter(
                empr_empr=empresa_id, 
                empr_codi=filial_id
            ).first()
            regime = filial.empr_regi_trib if filial else '1'
        except Exception:
            regime = '1'
        return get_csts_por_regime(regime)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        try:
            ctx['hoje'] = timezone.localdate()
        except Exception:
            ctx['hoje'] = timezone.now().date()
        ctx['lotes'] = []
        # Formset vazio por enquanto; será inicializado após salvar o produto
        from Produtos.models import Tabelaprecos
        ctx['formset'] = TabelaprecosFormSet(queryset=Tabelaprecos.objects.none(), prefix='precos')
        from Produtos.models import TabelaprecosPromocional
        ctx['promo_formset'] = TabelaprecosPromocionalFormSet(queryset=TabelaprecosPromocional.objects.none(), prefix='precos_promo')
        cst_choices = self._get_cst_choices()
        if self.request.POST:
            ctx['fiscal_form'] = ProdutoFiscalPadraoForm(self.request.POST, prefix='fiscal', cst_choices=cst_choices)
        else:
            ctx['fiscal_form'] = ProdutoFiscalPadraoForm(prefix='fiscal', cst_choices=cst_choices)
        return ctx

    def form_valid(self, form):
        instance = form.save(commit=False)
        try:
            unme_input = form.cleaned_data.get('prod_unme')
            from Produtos.models import UnidadeMedida
            if unme_input and not isinstance(unme_input, UnidadeMedida):
                code = str(unme_input).strip().upper()
                from core.utils import get_ncm_master_db
                unme = UnidadeMedida.objects.using(self.db_alias).filter(unid_codi=code).first()
                if not unme:
                    master_alias = get_ncm_master_db(self.db_alias)
                    master_unme = UnidadeMedida.objects.using(master_alias).filter(unid_codi=code).first()
                    desc = getattr(master_unme, 'unid_desc', code) if master_unme else code
                    unme = UnidadeMedida(unid_codi=code, unid_desc=desc)
                    unme.save(using=self.db_alias)
                instance.prod_unme = unme
        except Exception:
            pass
        try:
            unme_input = form.cleaned_data.get('prod_unme')
            if unme_input and not isinstance(unme_input, UnidadeMedida):
                code = str(unme_input).strip().upper()
                from core.utils import get_ncm_master_db
                # Busca no tenant
                unme = UnidadeMedida.objects.using(self.db_alias).filter(unid_codi=code).first()
                if not unme:
                    # Busca no master e clona
                    master_alias = get_ncm_master_db(self.db_alias)
                    master_unme = UnidadeMedida.objects.using(master_alias).filter(unid_codi=code).first()
                    desc = getattr(master_unme, 'unid_desc', code) if master_unme else code
                    unme = UnidadeMedida(unid_codi=code, unid_desc=desc)
                    unme.save(using=self.db_alias)
                instance.prod_unme = unme
        except Exception:
            pass
        # Atribuir empresa a partir de headers ou sessão
        empresa = (
            self.request.headers.get('X-Empresa')
            or self.request.META.get('HTTP_X_EMPRESA')
            or self.request.session.get('empresa_id')
            or self.empresa_id
        )
        if not empresa:
            form.add_error(None, 'Empresa não informada nos headers ou sessão.')
            return self.form_invalid(form)
        instance.prod_empr = str(empresa)
        # Garantir origem de mercadoria padrão como no serializer
        try:
            instance.prod_orig_merc = '0'
        except Exception:
            pass
        # Remover espaços e gerar código sequencial quando vazio
        if instance.prod_codi:
            instance.prod_codi = str(instance.prod_codi).strip()
        if not instance.prod_codi:
            # Geração sequencial por empresa, sem zeros à esquerda, evitando colisão
            ultimo = Produtos.objects.using(self.db_alias).filter(
                prod_empr=str(empresa)
            ).order_by('-prod_codi').first()
            proximo = int(ultimo.prod_codi) + 1 if (ultimo and str(ultimo.prod_codi).isdigit()) else 1
            while Produtos.objects.using(self.db_alias).filter(
                prod_empr=str(empresa), prod_codi=str(proximo)
            ).exists():
                proximo += 1
            instance.prod_codi = str(proximo)
        # Sincronizar prod_codi_nume quando disponível
        if hasattr(instance, 'prod_codi_nume'):
            instance.prod_codi_nume = instance.prod_codi
        # Tratar upload de foto (FileField externo ao ModelForm)
        uploaded = self.request.FILES.get('prod_foto') or form.cleaned_data.get('prod_foto')
        if uploaded:
            try:
                instance.prod_foto = uploaded.read()
            except Exception:
                pass
        try:
            with transaction.atomic(using=self.db_alias):
                instance.save(using=self.db_alias)
        except Exception as e:
            messages.error(self.request, f'Erro ao salvar produto: {e}')
            logger.exception('Falha ao salvar produto')
            return self.form_invalid(form)

        # Salvar preços (se enviados)
        formset = TabelaprecosPlainFormSet(self.request.POST, prefix='precos')
        promo_formset = TabelaprecosPromocionalPlainFormSet(self.request.POST, prefix='precos_promo')
        cst_choices = self._get_cst_choices()
        fiscal_form = ProdutoFiscalPadraoForm(self.request.POST, prefix='fiscal', cst_choices=cst_choices)

        if not formset.is_valid():
            for fs_form in formset.forms:
                for field, errs in fs_form.errors.items():
                    for err in errs:
                        messages.error(self.request, f'Preço - erro em {field}: {err}')
            ctx = self.get_context_data()
            ctx['form'] = form
            ctx['formset'] = formset
            ctx['promo_formset'] = promo_formset
            ctx['fiscal_form'] = fiscal_form
            return self.render_to_response(ctx)

        if not promo_formset.is_valid():
            for fs_form in promo_formset.forms:
                for field, errs in fs_form.errors.items():
                    for err in errs:
                        messages.error(self.request, f'Preço Promocional - erro em {field}: {err}')
            ctx = self.get_context_data()
            ctx['form'] = form
            ctx['formset'] = formset
            ctx['promo_formset'] = promo_formset
            ctx['fiscal_form'] = fiscal_form
            return self.render_to_response(ctx)
            
        if not fiscal_form.is_valid():
             for field, errs in fiscal_form.errors.items():
                for err in errs:
                    messages.error(self.request, f'Fiscal - erro em {field}: {err}')
             ctx = self.get_context_data()
             ctx['form'] = form
             ctx['formset'] = formset
             ctx['promo_formset'] = promo_formset
             ctx['fiscal_form'] = fiscal_form
             return self.render_to_response(ctx)

        logger.info(f"Preços POST: TOTAL_FORMS={formset.total_form_count()} INITIAL_FORMS={formset.initial_form_count()}")
        processed = 0
        processed_promo = 0
        try:
            with transaction.atomic(using=self.db_alias):
                # Salvar Fiscal Padrão
                fiscal_obj = fiscal_form.save(commit=False)
                # Verifica se algum campo foi preenchido
                has_fiscal_data = any(
                    getattr(fiscal_obj, field) is not None and getattr(fiscal_obj, field) != ''
                    for field in fiscal_form.fields
                )
                
                if has_fiscal_data:
                    fiscal_obj.produto = instance
                    try:
                         fiscal_obj.save(using=self.db_alias)
                    except Exception:
                         fiscal_obj.save()

                for f in formset.forms:
                    if not f.has_changed():
                        continue
                    cd = f.cleaned_data
                    if cd.get('DELETE'):
                        continue
                    tabe_prod = instance.prod_codi
                    try:
                        tabe_empr = int(self.empresa_id) if self.empresa_id else int(instance.prod_empr)
                    except Exception:
                        tabe_empr = self.empresa_id or instance.prod_empr
                    try:
                        tabe_fili = int(cd.get('tabe_fili')) if cd.get('tabe_fili') else (int(self.filial_id) if self.filial_id else 1)
                    except Exception:
                        tabe_fili = cd.get('tabe_fili') or (self.filial_id or 1)

                    existing = Tabelaprecos.objects.using(self.db_alias).filter(
                        tabe_empr=tabe_empr,
                        tabe_fili=tabe_fili,
                        tabe_prod=tabe_prod,
                    ).first()

                    if existing:
                        historico = "Alteração de preços via Web"
                        if cd.get('tabe_prco') is not None and existing.tabe_prco != cd.get('tabe_prco'):
                            historico += f"\nPreço Normal: R$ {float(existing.tabe_prco or 0):.2f} -> R$ {float(cd.get('tabe_prco') or 0):.2f}"
                        if cd.get('tabe_avis') is not None and existing.tabe_avis != cd.get('tabe_avis'):
                            historico += f"\nPreço à Vista: R$ {float(existing.tabe_avis or 0):.2f} -> R$ {float(cd.get('tabe_avis') or 0):.2f}"
                        if cd.get('tabe_apra') is not None and existing.tabe_apra != cd.get('tabe_apra'):
                            historico += f"\nPreço a Prazo: R$ {float(existing.tabe_apra or 0):.2f} -> R$ {float(cd.get('tabe_apra') or 0):.2f}"

                        hist_data = {
                            'tabe_empr': tabe_empr,
                            'tabe_fili': tabe_fili,
                            'tabe_prod': tabe_prod,
                            'tabe_data_hora': timezone.now(),
                            'tabe_hist': historico,
                            'tabe_perc_reaj': cd.get('tabe_perc_reaj'),
                            'tabe_prco_ante': existing.tabe_prco,
                            'tabe_avis_ante': existing.tabe_avis,
                            'tabe_apra_ante': existing.tabe_apra,
                            'tabe_pipi_ante': getattr(existing, 'tabe_pipi', None),
                            'tabe_fret_ante': getattr(existing, 'tabe_fret', None),
                            'tabe_desp_ante': getattr(existing, 'tabe_desp', None),
                            'tabe_cust_ante': getattr(existing, 'tabe_cust', None),
                            'tabe_cuge_ante': getattr(existing, 'tabe_cuge', None),
                            'tabe_icms_ante': getattr(existing, 'tabe_icms', None),
                            'tabe_impo_ante': getattr(existing, 'tabe_impo', None),
                            'tabe_marg_ante': getattr(existing, 'tabe_marg', None),
                            'tabe_praz_ante': getattr(existing, 'tabe_praz', None),
                            'tabe_valo_st_ante': getattr(existing, 'tabe_valo_st', None),
                            'tabe_prco_novo': cd.get('tabe_prco'),
                            'tabe_avis_novo': cd.get('tabe_avis'),
                            'tabe_apra_novo': cd.get('tabe_apra'),
                            'tabe_pipi_novo': cd.get('tabe_pipi'),
                            'tabe_fret_novo': cd.get('tabe_fret'),
                            'tabe_desp_novo': cd.get('tabe_desp'),
                            'tabe_cust_novo': cd.get('tabe_cust'),
                            'tabe_cuge_novo': cd.get('tabe_cuge'),
                            'tabe_icms_novo': cd.get('tabe_icms'),
                            'tabe_impo_novo': cd.get('tabe_impo'),
                            'tabe_marg_novo': cd.get('tabe_marg'),
                            'tabe_praz_novo': cd.get('tabe_praz'),
                            'tabe_valo_st_novo': cd.get('tabe_valo_st'),
                        }
                        Tabelaprecoshist.objects.using(self.db_alias).create(**hist_data)
                        rows = Tabelaprecos.objects.using(self.db_alias).filter(
                            tabe_empr=tabe_empr,
                            tabe_fili=tabe_fili,
                            tabe_prod=tabe_prod,
                        ).update(
                            tabe_prco=cd.get('tabe_prco'),
                            tabe_icms=cd.get('tabe_icms'),
                            tabe_desc=cd.get('tabe_desc'),
                            tabe_vipi=cd.get('tabe_vipi'),
                            tabe_pipi=cd.get('tabe_pipi'),
                            tabe_fret=cd.get('tabe_fret'),
                            tabe_desp=cd.get('tabe_desp'),
                            tabe_cust=cd.get('tabe_cust'),
                            tabe_marg=cd.get('tabe_marg'),
                            tabe_impo=cd.get('tabe_impo'),
                            tabe_avis=cd.get('tabe_avis'),
                            tabe_praz=cd.get('tabe_praz'),
                            tabe_apra=cd.get('tabe_apra'),
                            tabe_vare=cd.get('tabe_vare'),
                            field_log_data=cd.get('field_log_data'),
                            field_log_time=cd.get('field_log_time'),
                            tabe_valo_st=cd.get('tabe_valo_st'),
                            tabe_perc_reaj=cd.get('tabe_perc_reaj'),
                            tabe_hist=cd.get('tabe_hist'),
                            tabe_cuge=cd.get('tabe_cuge'),
                            tabe_entr=cd.get('tabe_entr'),
                            tabe_perc_st=cd.get('tabe_perc_st'),
                        )
                        logger.info(f"Atualização de preços: {rows} linha(s) para produto={tabe_prod} empr={tabe_empr} fili={tabe_fili}")
                    else:
                        hist_data = {
                            'tabe_empr': tabe_empr,
                            'tabe_fili': tabe_fili,
                            'tabe_prod': tabe_prod,
                            'tabe_data_hora': timezone.now(),
                            'tabe_hist': "Criação de preços via Web",
                            'tabe_perc_reaj': cd.get('tabe_perc_reaj'),
                            'tabe_prco_novo': cd.get('tabe_prco'),
                            'tabe_avis_novo': cd.get('tabe_avis'),
                            'tabe_apra_novo': cd.get('tabe_apra'),
                        }
                        Tabelaprecoshist.objects.using(self.db_alias).create(**hist_data)
                        obj = Tabelaprecos.objects.using(self.db_alias).create(
                            tabe_empr=tabe_empr,
                            tabe_fili=tabe_fili,
                            tabe_prod=tabe_prod,
                            tabe_prco=cd.get('tabe_prco'),
                            tabe_icms=cd.get('tabe_icms'),
                            tabe_desc=cd.get('tabe_desc'),
                            tabe_vipi=cd.get('tabe_vipi'),
                            tabe_pipi=cd.get('tabe_pipi'),
                            tabe_fret=cd.get('tabe_fret'),
                            tabe_desp=cd.get('tabe_desp'),
                            tabe_cust=cd.get('tabe_cust'),
                            tabe_marg=cd.get('tabe_marg'),
                            tabe_impo=cd.get('tabe_impo'),
                            tabe_avis=cd.get('tabe_avis'),
                            tabe_praz=cd.get('tabe_praz'),
                            tabe_apra=cd.get('tabe_apra'),
                            tabe_vare=cd.get('tabe_vare'),
                            field_log_data=cd.get('field_log_data'),
                            field_log_time=cd.get('field_log_time'),
                            tabe_valo_st=cd.get('tabe_valo_st'),
                            tabe_perc_reaj=cd.get('tabe_perc_reaj'),
                            tabe_hist=cd.get('tabe_hist'),
                            tabe_cuge=cd.get('tabe_cuge'),
                            tabe_entr=cd.get('tabe_entr'),
                            tabe_perc_st=cd.get('tabe_perc_st'),
                        )
                        logger.info(f"Criação de preços: produto={obj.tabe_prod} empr={obj.tabe_empr} fili={obj.tabe_fili}")
                    processed += 1

                for f in promo_formset.forms:
                    if not f.has_changed():
                        continue
                    cd = f.cleaned_data
                    if cd.get('DELETE'):
                        continue
                    tabe_prod = instance.prod_codi
                    try:
                        tabe_empr = int(self.empresa_id) if self.empresa_id else int(instance.prod_empr)
                    except Exception:
                        tabe_empr = self.empresa_id or instance.prod_empr
                    try:
                        tabe_fili = int(cd.get('tabe_fili')) if cd.get('tabe_fili') else (int(self.filial_id) if self.filial_id else 1)
                    except Exception:
                        tabe_fili = cd.get('tabe_fili') or (self.filial_id or 1)

                    existing_promo = TabelaprecosPromocional.objects.using(self.db_alias).filter(
                        tabe_empr=tabe_empr,
                        tabe_fili=tabe_fili,
                        tabe_prod=tabe_prod,
                    ).first()

                    dados_promo = {
                        'tabe_empr': tabe_empr,
                        'tabe_fili': tabe_fili,
                        'tabe_prod': tabe_prod,
                        'tabe_prco': cd.get('tabe_prco'),
                        'tabe_desp': cd.get('tabe_desp'),
                        'tabe_cust': cd.get('tabe_cust'),
                        'tabe_marg': cd.get('tabe_marg'),
                        'tabe_cuge': cd.get('tabe_cuge'),
                        'tabe_avis': cd.get('tabe_avis'),
                        'tabe_praz': cd.get('tabe_praz'),
                        'tabe_apra': cd.get('tabe_apra'),
                        'tabe_hist': cd.get('tabe_hist'),
                        'tabe_perc_reaj': cd.get('tabe_perc_reaj'),
                    }

                    if existing_promo:
                        atualizar_preco_promocional_com_historico(self.db_alias, existing_promo, dados_promo)
                    else:
                        criar_preco_promocional_com_historico(self.db_alias, dados_promo)
                    processed_promo += 1
        except Exception as e:
            messages.error(self.request, f'Erro ao salvar preços: {e}')
            logger.exception('Falha ao salvar tabela de preços')
            ctx = self.get_context_data()
            ctx['form'] = form
            ctx['formset'] = formset
            ctx['promo_formset'] = promo_formset
            return self.render_to_response(ctx)
        if processed == 0 and processed_promo == 0:
            messages.info(self.request, 'Nenhum item de preço foi enviado/alterado.')
        messages.success(self.request, f'Produto criado com sucesso. Código: {instance.prod_codi}')
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        # Consolidar erros do formulário principal em mensagens para o usuário
        if form.errors:
            for field, errs in form.errors.items():
                for err in errs:
                    if field == '__all__':
                        messages.error(self.request, f'{err}')
                    else:
                        messages.error(self.request, f'Erro em {field}: {err}')
        # Também validar o formset de preços, se presente na requisição
        from ...models import Tabelaprecos
        try:
            formset = TabelaprecosFormSet(self.request.POST, queryset=Tabelaprecos.objects.none(), prefix='precos')
            if not formset.is_valid():
                for fs_form in formset.forms:
                    if fs_form.errors:
                        for field, errs in fs_form.errors.items():
                            for err in errs:
                                messages.error(self.request, f'Preço - erro em {field}: {err}')
        except Exception:
            pass
        from ...models import TabelaprecosPromocional
        try:
            promo_formset = TabelaprecosPromocionalFormSet(self.request.POST, queryset=TabelaprecosPromocional.objects.none(), prefix='precos_promo')
            if not promo_formset.is_valid():
                for fs_form in promo_formset.forms:
                    if fs_form.errors:
                        for field, errs in fs_form.errors.items():
                            for err in errs:
                                messages.error(self.request, f'Preço Promocional - erro em {field}: {err}')
        except Exception:
            pass
        return super().form_invalid(form)


class ProdutoUpdateView(DBAndSlugMixin, UpdateView):
    model = Produtos
    form_class = ProdutosForm
    template_name = 'Produtos/produtos_update.html'
    pk_url_kwarg = 'prod_codi'
    slug_url_kwarg = 'slug'

    def _get_cst_choices(self):
        try:
            empresa_id = int(self.empresa_id or 1)
            filial_id = int(self.filial_id or 1)
            filial = Filiais.objects.using(self.db_alias).filter(
                empr_empr=empresa_id, 
                empr_codi=filial_id
            ).first()
            regime = filial.empr_regi_trib if filial else '1'
        except Exception:
            regime = '1'
        return get_csts_por_regime(regime)


    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        if 'prod_codi' in form.fields:
            form.fields['prod_codi'].disabled = True
            form.fields['prod_codi'].required = False
        return form

    def get_object(self, queryset=None):
        prod_codi = self.kwargs.get('prod_codi')
        if not prod_codi:
            raise Http404('Código do produto não informado')
        # Buscar por código e, se informado, por empresa para evitar múltiplos resultados
        qs = Produtos.objects.using(self.db_alias).filter(prod_codi=prod_codi)
        if self.empresa_id:
            qs = qs.filter(prod_empr=str(self.empresa_id))
        obj = qs.order_by('prod_empr').first()
        if not obj:
            raise Http404('Produto não encontrado')
        return obj

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        try:
            ctx['hoje'] = timezone.localdate()
        except Exception:
            ctx['hoje'] = timezone.now().date()
        # Carregar preços vinculados ao produto via tabe_prod/tabe_empr
        produto = self.object
        try:
            lote_empr = int(produto.prod_empr)
        except Exception:
            lote_empr = produto.prod_empr
        try:
            ctx['lotes'] = list(
                Lote.objects.using(self.db_alias)
                .filter(lote_empr=lote_empr, lote_prod=str(produto.prod_codi))
                .order_by('-lote_lote')
            )
        except Exception:
            ctx['lotes'] = []
        try:
            from django.db.models import Sum
            from decimal import Decimal
            from Produtos.models import SaldoProduto
            lotes_total = (
                Lote.objects.using(self.db_alias)
                .filter(lote_empr=lote_empr, lote_prod=str(produto.prod_codi), lote_ativ=True)
                .aggregate(total=Sum('lote_sald'))
                .get('total')
            )
            saldo_lotes = Decimal(str(lotes_total or 0))
            filial_id = self.filial_id or self.request.session.get('filial_id') or 1
            sp = (
                SaldoProduto.objects.using(self.db_alias)
                .filter(produto_codigo=produto, empresa=str(produto.prod_empr), filial=str(filial_id))
                .first()
            )
            saldo_total = Decimal(str(getattr(sp, 'saldo_estoque', 0) or 0))
            saldo_sem_lote = saldo_total - saldo_lotes
            ctx['saldo_total'] = saldo_total
            ctx['saldo_lotes'] = saldo_lotes
            ctx['saldo_sem_lote'] = saldo_sem_lote
        except Exception:
            ctx['saldo_total'] = None
            ctx['saldo_lotes'] = None
            ctx['saldo_sem_lote'] = None
        try:
            emp_int = int(produto.prod_empr)
        except Exception:
            emp_int = produto.prod_empr
        qs = Tabelaprecos.objects.using(self.db_alias).filter(
            tabe_prod=produto.prod_codi,
            tabe_empr=emp_int
        )
        initial_list = []
        for preco in qs:
            initial_list.append({
                'tabe_fili': preco.tabe_fili,
                'tabe_prco': preco.tabe_prco,
                'tabe_icms': preco.tabe_icms,
                'tabe_desc': preco.tabe_desc,
                'tabe_vipi': preco.tabe_vipi,
                'tabe_pipi': preco.tabe_pipi,
                'tabe_fret': preco.tabe_fret,
                'tabe_desp': preco.tabe_desp,
                'tabe_cust': preco.tabe_cust,
                'tabe_marg': preco.tabe_marg,
                'tabe_impo': preco.tabe_impo,
                'tabe_avis': preco.tabe_avis,
                'tabe_praz': preco.tabe_praz,
                'tabe_apra': preco.tabe_apra,
                'tabe_vare': getattr(preco, 'tabe_vare', None),
                'tabe_hist': getattr(preco, 'tabe_hist', None),
                'tabe_cuge': getattr(preco, 'tabe_cuge', None),
                'tabe_entr': getattr(preco, 'tabe_entr', None),
                'tabe_perc_st': getattr(preco, 'tabe_perc_st', None),
            })
        if not initial_list:
            try:
                default_fili = int(self.filial_id or self.request.session.get('filial_id') or 1)
            except Exception:
                default_fili = self.filial_id or self.request.session.get('filial_id') or 1
            initial_list = [{'tabe_fili': default_fili}]
        ctx['formset'] = TabelaprecosPlainFormSet(initial=initial_list, prefix='precos')

        qs_promo = TabelaprecosPromocional.objects.using(self.db_alias).filter(
            tabe_prod=produto.prod_codi,
            tabe_empr=emp_int,
        )
        promo_initial_list = []
        for preco in qs_promo:
            promo_initial_list.append({
                'tabe_fili': preco.tabe_fili,
                'tabe_prco': preco.tabe_prco,
                'tabe_desp': preco.tabe_desp,
                'tabe_cust': preco.tabe_cust,
                'tabe_marg': preco.tabe_marg,
                'tabe_cuge': preco.tabe_cuge,
                'tabe_avis': preco.tabe_avis,
                'tabe_praz': preco.tabe_praz,
                'tabe_apra': preco.tabe_apra,
                'tabe_hist': preco.tabe_hist,
                'tabe_perc_reaj': preco.tabe_perc_reaj,
            })
        if not promo_initial_list:
            try:
                default_fili = int(self.filial_id or self.request.session.get('filial_id') or 1)
            except Exception:
                default_fili = self.filial_id or self.request.session.get('filial_id') or 1
            promo_initial_list = [{'tabe_fili': default_fili}]
        ctx['promo_formset'] = TabelaprecosPromocionalPlainFormSet(initial=promo_initial_list, prefix='precos_promo')
        try:
            ctx['promo_historico'] = list(
                TabelaprecosPromocionalhist.objects.using(self.db_alias).filter(
                    tabe_prod=produto.prod_codi,
                    tabe_empr=emp_int,
                ).order_by('-tabe_data_hora')[:50]
            )
        except Exception:
            ctx['promo_historico'] = []

        # Load Fiscal Data
        try:
            fiscal_obj = ProdutoFiscalPadrao.objects.using(self.db_alias).filter(produto=produto).first()
        except Exception:
            fiscal_obj = None
        
        cst_choices = self._get_cst_choices()
        if self.request.POST:
             ctx['fiscal_form'] = ProdutoFiscalPadraoForm(self.request.POST, prefix='fiscal', instance=fiscal_obj, cst_choices=cst_choices)
        else:
             ctx['fiscal_form'] = ProdutoFiscalPadraoForm(prefix='fiscal', instance=fiscal_obj, cst_choices=cst_choices)

        logger.info(f'Preços carregados para produto {produto.prod_codi}: {initial_list}')
        return ctx

    def form_valid(self, form):
        instance = form.save(commit=False)
        # Garantir que o código (PK) permaneça o mesmo do objeto carregado
        instance.prod_codi = form.instance.prod_codi
        # Manter/atualizar empresa a partir de headers ou sessão, quando presente
        empresa = (
            self.request.headers.get('X-Empresa')
            or self.request.META.get('HTTP_X_EMPRESA')
            or self.request.session.get('empresa_id')
            or self.empresa_id
        )
        if empresa:
            instance.prod_empr = str(empresa)
        # Tratar upload de foto (FileField externo ao ModelForm)
        uploaded = self.request.FILES.get('prod_foto') or form.cleaned_data.get('prod_foto')
        if uploaded:
            try:
                instance.prod_foto = uploaded.read()
            except Exception:
                pass
        update_fields = [
            'prod_nome','prod_unme','prod_grup','prod_sugr','prod_fami',
            'prod_loca','prod_ncm','prod_gtin','prod_marc','prod_orig_merc',
            'prod_cera_m2cx','prod_cera_pccx','prod_cera_kgcx','prod_cera_m2pallet',
            'prod_cera_form','prod_cera_espe','prod_cera_cor','prod_cera_cole',
            'prod_cera_tipo','prod_cera_esti',
        ]
        if uploaded:
            update_fields.append('prod_foto')
        instance.save(using=self.db_alias, update_fields=update_fields)

        # Atualizar preços
        try:
            emp_int = int(self.empresa_id) if self.empresa_id else int(form.instance.prod_empr)
        except Exception:
            emp_int = self.empresa_id or form.instance.prod_empr
        formset = TabelaprecosPlainFormSet(self.request.POST, prefix='precos')
        promo_formset = TabelaprecosPromocionalPlainFormSet(self.request.POST, prefix='precos_promo')

        # Load Fiscal Data for saving
        try:
             fiscal_obj = ProdutoFiscalPadrao.objects.using(self.db_alias).filter(produto=instance).first()
        except Exception:
             fiscal_obj = None
        cst_choices = self._get_cst_choices()
        fiscal_form = ProdutoFiscalPadraoForm(self.request.POST, prefix='fiscal', instance=fiscal_obj, cst_choices=cst_choices)

        if not formset.is_valid():
            for fs_form in formset.forms:
                for field, errs in fs_form.errors.items():
                    for err in errs:
                        messages.error(self.request, f'Preço - erro em {field}: {err}')
            ctx = self.get_context_data()
            ctx['form'] = form
            ctx['formset'] = formset
            ctx['promo_formset'] = promo_formset
            ctx['fiscal_form'] = fiscal_form
            return self.render_to_response(ctx)

        if not promo_formset.is_valid():
            for fs_form in promo_formset.forms:
                for field, errs in fs_form.errors.items():
                    for err in errs:
                        messages.error(self.request, f'Preço Promocional - erro em {field}: {err}')
            ctx = self.get_context_data()
            ctx['form'] = form
            ctx['formset'] = formset
            ctx['promo_formset'] = promo_formset
            ctx['fiscal_form'] = fiscal_form
            return self.render_to_response(ctx)
            
        if not fiscal_form.is_valid():
            for field, errs in fiscal_form.errors.items():
                for err in errs:
                    messages.error(self.request, f'Fiscal - erro em {field}: {err}')
            ctx = self.get_context_data()
            ctx['form'] = form
            ctx['formset'] = formset
            ctx['promo_formset'] = promo_formset
            ctx['fiscal_form'] = fiscal_form
            return self.render_to_response(ctx)

        # Save Fiscal Data
        fiscal_obj_save = fiscal_form.save(commit=False)
        has_fiscal_data = any(
             getattr(fiscal_obj_save, field) is not None and getattr(fiscal_obj_save, field) != ''
             for field in fiscal_form.fields
        )
        if has_fiscal_data:
             fiscal_obj_save.produto = instance
             try:
                  fiscal_obj_save.save(using=self.db_alias)
             except Exception:
                  fiscal_obj_save.save()
        elif fiscal_obj:
             # If it existed but now is empty (unlikely with required=False but possible if fields cleared), maybe delete? 
             # For now, we just save what we have. If all empty, it just updates to empty.
             try:
                  fiscal_obj_save.save(using=self.db_alias)
             except Exception:
                  fiscal_obj_save.save()

        processed = 0
        processed_promo = 0
        for f in formset.forms:
            if not f.has_changed():
                continue
            cd = f.cleaned_data
            if cd.get('DELETE'):
                continue
            tabe_prod = instance.prod_codi
            try:
                tabe_empr = int(self.empresa_id) if self.empresa_id else int(instance.prod_empr)
            except Exception:
                tabe_empr = self.empresa_id or instance.prod_empr
            try:
                tabe_fili = int(cd.get('tabe_fili')) if cd.get('tabe_fili') else (int(self.filial_id) if self.filial_id else 1)
            except Exception:
                tabe_fili = cd.get('tabe_fili') or (self.filial_id or 1)

            existing = Tabelaprecos.objects.using(self.db_alias).filter(
                tabe_empr=tabe_empr,
                tabe_fili=tabe_fili,
                tabe_prod=tabe_prod,
            ).first()

            if existing:
                historico = "Alteração de preços via Web"
                if cd.get('tabe_prco') is not None and existing.tabe_prco != cd.get('tabe_prco'):
                    historico += f"\nPreço Normal: R$ {float(existing.tabe_prco or 0):.2f} -> R$ {float(cd.get('tabe_prco') or 0):.2f}"
                if cd.get('tabe_avis') is not None and existing.tabe_avis != cd.get('tabe_avis'):
                    historico += f"\nPreço à Vista: R$ {float(existing.tabe_avis or 0):.2f} -> R$ {float(cd.get('tabe_avis') or 0):.2f}"
                if cd.get('tabe_apra') is not None and existing.tabe_apra != cd.get('tabe_apra'):
                    historico += f"\nPreço a Prazo: R$ {float(existing.tabe_apra or 0):.2f} -> R$ {float(cd.get('tabe_apra') or 0):.2f}"

                hist_data = {
                    'tabe_empr': tabe_empr,
                    'tabe_fili': tabe_fili,
                    'tabe_prod': tabe_prod,
                    'tabe_data_hora': timezone.now(),
                    'tabe_hist': historico,
                    'tabe_perc_reaj': cd.get('tabe_perc_reaj'),
                    'tabe_prco_ante': existing.tabe_prco,
                    'tabe_avis_ante': existing.tabe_avis,
                    'tabe_apra_ante': existing.tabe_apra,
                    'tabe_pipi_ante': getattr(existing, 'tabe_pipi', None),
                    'tabe_fret_ante': getattr(existing, 'tabe_fret', None),
                    'tabe_desp_ante': getattr(existing, 'tabe_desp', None),
                    'tabe_cust_ante': getattr(existing, 'tabe_cust', None),
                    'tabe_cuge_ante': getattr(existing, 'tabe_cuge', None),
                    'tabe_icms_ante': getattr(existing, 'tabe_icms', None),
                    'tabe_impo_ante': getattr(existing, 'tabe_impo', None),
                    'tabe_marg_ante': getattr(existing, 'tabe_marg', None),
                    'tabe_praz_ante': getattr(existing, 'tabe_praz', None),
                    'tabe_valo_st_ante': getattr(existing, 'tabe_valo_st', None),
                    'tabe_prco_novo': cd.get('tabe_prco'),
                    'tabe_avis_novo': cd.get('tabe_avis'),
                    'tabe_apra_novo': cd.get('tabe_apra'),
                    'tabe_pipi_novo': cd.get('tabe_pipi'),
                    'tabe_fret_novo': cd.get('tabe_fret'),
                    'tabe_desp_novo': cd.get('tabe_desp'),
                    'tabe_cust_novo': cd.get('tabe_cust'),
                    'tabe_cuge_novo': cd.get('tabe_cuge'),
                    'tabe_icms_novo': cd.get('tabe_icms'),
                    'tabe_impo_novo': cd.get('tabe_impo'),
                    'tabe_marg_novo': cd.get('tabe_marg'),
                    'tabe_praz_novo': cd.get('tabe_praz'),
                    'tabe_valo_st_novo': cd.get('tabe_valo_st'),
                }
                Tabelaprecoshist.objects.using(self.db_alias).create(**hist_data)
                Tabelaprecos.objects.using(self.db_alias).filter(
                    tabe_empr=tabe_empr,
                    tabe_fili=tabe_fili,
                    tabe_prod=tabe_prod,
                ).update(
                    tabe_prco=cd.get('tabe_prco'),
                    tabe_icms=cd.get('tabe_icms'),
                    tabe_desc=cd.get('tabe_desc'),
                    tabe_vipi=cd.get('tabe_vipi'),
                    tabe_pipi=cd.get('tabe_pipi'),
                    tabe_fret=cd.get('tabe_fret'),
                    tabe_desp=cd.get('tabe_desp'),
                    tabe_cust=cd.get('tabe_cust'),
                    tabe_marg=cd.get('tabe_marg'),
                    tabe_impo=cd.get('tabe_impo'),
                    tabe_avis=cd.get('tabe_avis'),
                    tabe_praz=cd.get('tabe_praz'),
                    tabe_apra=cd.get('tabe_apra'),
                    tabe_vare=cd.get('tabe_vare'),
                    field_log_data=cd.get('field_log_data'),
                    field_log_time=cd.get('field_log_time'),
                    tabe_valo_st=cd.get('tabe_valo_st'),
                    tabe_perc_reaj=cd.get('tabe_perc_reaj'),
                    tabe_hist=cd.get('tabe_hist'),
                    tabe_cuge=cd.get('tabe_cuge'),
                    tabe_entr=cd.get('tabe_entr'),
                    tabe_perc_st=cd.get('tabe_perc_st'),
                )
            else:
                hist_data = {
                    'tabe_empr': tabe_empr,
                    'tabe_fili': tabe_fili,
                    'tabe_prod': tabe_prod,
                    'tabe_data_hora': timezone.now(),
                    'tabe_hist': "Criação de preços via Web",
                    'tabe_perc_reaj': cd.get('tabe_perc_reaj'),
                    'tabe_prco_novo': cd.get('tabe_prco'),
                    'tabe_avis_novo': cd.get('tabe_avis'),
                    'tabe_apra_novo': cd.get('tabe_apra'),
                }
                Tabelaprecoshist.objects.using(self.db_alias).create(**hist_data)
                Tabelaprecos.objects.using(self.db_alias).create(
                    tabe_empr=tabe_empr,
                    tabe_fili=tabe_fili,
                    tabe_prod=tabe_prod,
                    tabe_prco=cd.get('tabe_prco'),
                    tabe_icms=cd.get('tabe_icms'),
                    tabe_desc=cd.get('tabe_desc'),
                    tabe_vipi=cd.get('tabe_vipi'),
                    tabe_pipi=cd.get('tabe_pipi'),
                    tabe_fret=cd.get('tabe_fret'),
                    tabe_desp=cd.get('tabe_desp'),
                    tabe_cust=cd.get('tabe_cust'),
                    tabe_marg=cd.get('tabe_marg'),
                    tabe_impo=cd.get('tabe_impo'),
                    tabe_avis=cd.get('tabe_avis'),
                    tabe_praz=cd.get('tabe_praz'),
                    tabe_apra=cd.get('tabe_apra'),
                    tabe_vare=cd.get('tabe_vare'),
                    field_log_data=cd.get('field_log_data'),
                    field_log_time=cd.get('field_log_time'),
                    tabe_valo_st=cd.get('tabe_valo_st'),
                    tabe_perc_reaj=cd.get('tabe_perc_reaj'),
                    tabe_hist=cd.get('tabe_hist'),
                    tabe_cuge=cd.get('tabe_cuge'),
                    tabe_entr=cd.get('tabe_entr'),
                    tabe_perc_st=cd.get('tabe_perc_st'),
                )
            processed += 1

        for f in promo_formset.forms:
            if not f.has_changed():
                continue
            cd = f.cleaned_data
            if cd.get('DELETE'):
                continue
            tabe_prod = instance.prod_codi
            try:
                tabe_empr = int(self.empresa_id) if self.empresa_id else int(instance.prod_empr)
            except Exception:
                tabe_empr = self.empresa_id or instance.prod_empr
            try:
                tabe_fili = int(cd.get('tabe_fili')) if cd.get('tabe_fili') else (int(self.filial_id) if self.filial_id else 1)
            except Exception:
                tabe_fili = cd.get('tabe_fili') or (self.filial_id or 1)

            existing_promo = TabelaprecosPromocional.objects.using(self.db_alias).filter(
                tabe_empr=tabe_empr,
                tabe_fili=tabe_fili,
                tabe_prod=tabe_prod,
            ).first()

            dados_promo = {
                'tabe_empr': tabe_empr,
                'tabe_fili': tabe_fili,
                'tabe_prod': tabe_prod,
                'tabe_prco': cd.get('tabe_prco'),
                'tabe_desp': cd.get('tabe_desp'),
                'tabe_cust': cd.get('tabe_cust'),
                'tabe_marg': cd.get('tabe_marg'),
                'tabe_cuge': cd.get('tabe_cuge'),
                'tabe_avis': cd.get('tabe_avis'),
                'tabe_praz': cd.get('tabe_praz'),
                'tabe_apra': cd.get('tabe_apra'),
                'tabe_hist': cd.get('tabe_hist'),
                'tabe_perc_reaj': cd.get('tabe_perc_reaj'),
            }

            if existing_promo:
                atualizar_preco_promocional_com_historico(self.db_alias, existing_promo, dados_promo)
            else:
                criar_preco_promocional_com_historico(self.db_alias, dados_promo)
            processed_promo += 1

        if processed == 0 and processed_promo == 0:
            messages.info(self.request, 'Nenhum item de preço foi enviado/alterado.')
        messages.success(self.request, f'Produto atualizado com sucesso. Código: {instance.prod_codi}')
        logger.info(f'Produto atualizado com sucesso. Código: {instance.prod_codi},tabela de preços atualizada com sucesso.')
        return redirect(self.get_success_url())

# Unidades de Medida
class UnidadeMedidaListView(DBAndSlugMixin, ListView):
    model = UnidadeMedida
    template_name = 'Produtos/unidades_list.html'
    context_object_name = 'unidades'
    paginate_by = 20
    
    
    def get_queryset(self):
        qs = UnidadeMedida.objects.using(self.db_alias).all()
        descricao = (self.request.GET.get('descricao') or '').strip()
        if descricao:
            qs = qs.filter(unid_desc__icontains=descricao)
        return qs
    
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        ctx['descricao'] = (self.request.GET.get('descricao') or '').strip()
        return ctx

class UnidadeMedidaCreateView(DBAndSlugMixin, CreateView):
    model = UnidadeMedida
    form_class = UnidadeMedidaForm
    template_name = 'Produtos/unidade_create.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        return ctx
    
    def form_valid(self, form):
        obj = form.save(commit=False)
        # O código (unid_codi) é inserido manualmente pelo usuário
        if obj.unid_codi:
            obj.unid_codi = obj.unid_codi.upper()
        obj.save(using=self.db_alias)
        messages.success(self.request, 'Unidade de medida criada com sucesso.')
        return redirect(self.get_success_url())


class UnidadeMedidaUpdateView(DBAndSlugMixin, UpdateView):
    model = UnidadeMedida
    form_class = UnidadeMedidaForm
    template_name = 'Produtos/unidade_update.html'
    pk_url_kwarg = 'codigo'

    def get_queryset(self):
        return UnidadeMedida.objects.using(self.db_alias).all()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        return ctx

    def form_valid(self, form):
        obj = form.save(commit=False)
        if obj.unid_codi:
            obj.unid_codi = obj.unid_codi.upper()
        obj.save(using=self.db_alias)
        messages.success(self.request, 'Unidade de medida atualizada com sucesso.')
        return redirect(self.get_success_url())


class UnidadeMedidaDeleteView(DBAndSlugMixin, DeleteView):
    model = UnidadeMedida
    template_name = 'Produtos/unidade_delete.html'
    pk_url_kwarg = 'codigo'
    success_url = reverse_lazy('unidades_list')

    def get_queryset(self):
        return UnidadeMedida.objects.using(self.db_alias).all()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        return ctx

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.save(using=self.db_alias)
        messages.success(self.request, 'Unidade de medida excluída com sucesso.')
        return redirect(self.get_success_url())

# Grupos
class GrupoListView(DBAndSlugMixin, ListView):
    model = GrupoProduto
    template_name = 'Produtos/grupos_list.html'
    context_object_name = 'grupos'
    paginate_by = 20

    def get_queryset(self):
        qs = GrupoProduto.objects.using(self.db_alias).all()
        nome = (self.request.GET.get('nome') or '').strip()
        if nome:
            qs = qs.filter(descricao__icontains=nome)
        return qs.order_by('codigo')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        ctx['nome'] = (self.request.GET.get('nome') or '').strip()
        return ctx

class GrupoCreateView(DBAndSlugMixin, CreateView):
    model = GrupoProduto
    form_class = GrupoForm
    template_name = 'Produtos/grupo_create.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        return ctx

    def form_valid(self, form):
        obj = form.save(commit=False)
        from django.db.models import IntegerField
        from django.db.models.functions import Cast
        qs = GrupoProduto.objects.using(self.db_alias)
        maior = qs.annotate(codigo_int=Cast('codigo', IntegerField())).order_by('-codigo_int').first()
        proximo = (int(maior.codigo) + 1) if maior else 1
        obj.codigo = str(proximo)
        obj.save(using=self.db_alias)
        messages.success(self.request, 'Grupo criado com sucesso.')
        return redirect(self.get_success_url())

class GrupoUpdateView(DBAndSlugMixin, UpdateView):
    model = GrupoProduto
    form_class = GrupoForm
    template_name = 'Produtos/grupo_update.html'
    pk_url_kwarg = 'codigo'

    def get_queryset(self):
        return GrupoProduto.objects.using(self.db_alias).all()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        return ctx

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.save(using=self.db_alias)
        messages.success(self.request, 'Grupo atualizado com sucesso.')
        return redirect(self.get_success_url())

class GrupoDeleteView(DBAndSlugMixin, DeleteView):
    model = GrupoProduto
    template_name = 'Produtos/grupo_delete.html'
    pk_url_kwarg = 'codigo'

    def get_queryset(self):
        return GrupoProduto.objects.using(self.db_alias).all()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        return ctx

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.delete(using=self.db_alias)
        messages.success(self.request, 'Grupo excluído com sucesso.')
        return redirect(self.get_success_url())

# Subgrupos
class SubgrupoListView(DBAndSlugMixin, ListView):
    model = SubgrupoProduto
    template_name = 'Produtos/subgrupos_list.html'
    context_object_name = 'subgrupos'
    paginate_by = 20

    def get_queryset(self):
        qs = SubgrupoProduto.objects.using(self.db_alias).all()
        nome = (self.request.GET.get('nome') or '').strip()
        if nome:
            qs = qs.filter(descricao__icontains=nome)
        return qs.order_by('codigo')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        ctx['nome'] = (self.request.GET.get('nome') or '').strip()
        return ctx

class SubgrupoCreateView(DBAndSlugMixin, CreateView):
    model = SubgrupoProduto
    form_class = SubgrupoForm
    template_name = 'Produtos/subgrupo_create.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        return ctx

    def form_valid(self, form):
        obj = form.save(commit=False)
        from django.db.models import IntegerField
        from django.db.models.functions import Cast
        qs = SubgrupoProduto.objects.using(self.db_alias)
        maior = qs.annotate(codigo_int=Cast('codigo', IntegerField())).order_by('-codigo_int').first()
        proximo = (int(maior.codigo) + 1) if maior else 1
        obj.codigo = str(proximo)
        obj.save(using=self.db_alias)
        messages.success(self.request, 'Subgrupo criado com sucesso.')
        return redirect(self.get_success_url())

class SubgrupoUpdateView(DBAndSlugMixin, UpdateView):
    model = SubgrupoProduto
    form_class = SubgrupoForm
    template_name = 'Produtos/subgrupo_update.html'
    pk_url_kwarg = 'codigo'

    def get_queryset(self):
        return SubgrupoProduto.objects.using(self.db_alias).all()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        return ctx

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.save(using=self.db_alias)
        messages.success(self.request, 'Subgrupo atualizado com sucesso.')
        return redirect(self.get_success_url())

class SubgrupoDeleteView(DBAndSlugMixin, DeleteView):
    model = SubgrupoProduto
    template_name = 'Produtos/subgrupo_delete.html'
    pk_url_kwarg = 'codigo'

    def get_queryset(self):
        return SubgrupoProduto.objects.using(self.db_alias).all()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        return ctx

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.delete(using=self.db_alias)
        messages.success(self.request, 'Subgrupo excluído com sucesso.')
        return redirect(self.get_success_url())

# Famílias
class FamiliaListView(DBAndSlugMixin, ListView):
    model = FamiliaProduto
    template_name = 'Produtos/familias_produto_list.html'
    context_object_name = 'familias'
    paginate_by = 20

    def get_queryset(self):
        qs = FamiliaProduto.objects.using(self.db_alias).all()
        nome = (self.request.GET.get('nome') or '').strip()
        if nome:
            qs = qs.filter(descricao__icontains=nome)
        return qs.order_by('codigo')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        ctx['nome'] = (self.request.GET.get('nome') or '').strip()
        return ctx

class FamiliaCreateView(DBAndSlugMixin, CreateView):
    model = FamiliaProduto
    form_class = FamiliaForm
    template_name = 'Produtos/familia_produto_create.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        return ctx

    def form_valid(self, form):
        obj = form.save(commit=False)
        from django.db.models import IntegerField
        from django.db.models.functions import Cast
        qs = FamiliaProduto.objects.using(self.db_alias)
        maior = qs.annotate(codigo_int=Cast('codigo', IntegerField())).order_by('-codigo_int').first()
        proximo = (int(maior.codigo) + 1) if maior else 1
        obj.codigo = str(proximo)
        obj.save(using=self.db_alias)
        messages.success(self.request, 'Família criada com sucesso.')
        return redirect(self.get_success_url())

class FamiliaUpdateView(DBAndSlugMixin, UpdateView):
    model = FamiliaProduto
    form_class = FamiliaForm
    template_name = 'Produtos/familia_produto_update.html'
    pk_url_kwarg = 'codigo'

    def get_queryset(self):
        return FamiliaProduto.objects.using(self.db_alias).all()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        return ctx

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.save(using=self.db_alias)
        messages.success(self.request, 'Família atualizada com sucesso.')
        return redirect(self.get_success_url())

class FamiliaDeleteView(DBAndSlugMixin, DeleteView):
    model = FamiliaProduto
    template_name = 'Produtos/familia_produto_delete.html'
    pk_url_kwarg = 'codigo'

    def get_queryset(self):
        return FamiliaProduto.objects.using(self.db_alias).all()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        return ctx

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.delete(using=self.db_alias)
        messages.success(self.request, 'Família excluída com sucesso.')
        return redirect(self.get_success_url())

# Marcas
class MarcaListViewWeb(DBAndSlugMixin, ListView):
    model = Marca
    template_name = 'Produtos/marcas_list.html'
    context_object_name = 'marcas'
    paginate_by = 20

    def get_queryset(self):
        qs = Marca.objects.using(self.db_alias).all()
        nome = (self.request.GET.get('nome') or '').strip()
        if nome:
            qs = qs.filter(nome__icontains=nome)
        return qs.order_by('codigo')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        ctx['nome'] = (self.request.GET.get('nome') or '').strip()
        return ctx

class MarcaCreateView(DBAndSlugMixin, CreateView):
    model = Marca
    form_class = MarcaForm
    template_name = 'Produtos/marca_create.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        return ctx

    def form_valid(self, form):
        obj = form.save(commit=False)
        maior_codigo = Marca.objects.using(self.db_alias).order_by('-codigo').first()
        if maior_codigo:
            obj.codigo = maior_codigo.codigo + 1
        else:
            obj.codigo = 1
        obj.save(using=self.db_alias)
        messages.success(self.request, 'Marca criada com sucesso.')
        return redirect(self.get_success_url())

class MarcaUpdateView(DBAndSlugMixin, UpdateView):
    model = Marca
    form_class = MarcaForm
    template_name = 'Produtos/marca_update.html'
    pk_url_kwarg = 'codigo'

    def get_queryset(self):
        return Marca.objects.using(self.db_alias).all()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        return ctx

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.save(using=self.db_alias)
        messages.success(self.request, 'Marca atualizada com sucesso.')
        return redirect(self.get_success_url())

class MarcaDeleteView(DBAndSlugMixin, DeleteView):
    model = Marca
    template_name = 'Produtos/marca_delete.html'
    pk_url_kwarg = 'codigo'

    def get_queryset(self):
        return Marca.objects.using(self.db_alias).all()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        return ctx

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.delete(using=self.db_alias)
        messages.success(self.request, 'Marca excluída com sucesso.')
        return redirect(self.get_success_url())

  

class ProdutoDeleteView(DBAndSlugMixin, DeleteView):
    model = Produtos
    template_name = 'Produtos/produto_confirm_delete.html'
    pk_url_kwarg = 'prod_codi'

    def get_object(self, queryset=None):
        prod_codi = self.kwargs.get('prod_codi')
        if not prod_codi:
            raise Http404('Código do produto não informado')
        qs = Produtos.objects.using(self.db_alias).filter(prod_codi=prod_codi)
        if self.empresa_id:
            qs = qs.filter(prod_empr=str(self.empresa_id))
        obj = qs.order_by('prod_empr').first()
        if not obj:
            raise Http404('Produto não encontrado')
        return obj

    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.delete(using=self.db_alias)
        messages.success(self.request, 'Produto excluído com sucesso.')
        return redirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        return ctx


class ExportarProdutosView(DBAndSlugMixin, View):
    def get(self, request, *args, **kwargs):
        prod_nome = (request.GET.get('prod_nome') or '').strip()
        prod_codi = (request.GET.get('prod_codi') or '').strip()

        qs = (
            Produtos.objects.using(self.db_alias)
            .all()
            .select_related(
                'prod_unme',
                'prod_grup',
                'prod_sugr',
                'prod_fami',
                'prod_marc',
            )
        )

        if self.empresa_id:
            qs = qs.filter(prod_empr=str(self.empresa_id))

        if prod_nome:
            qs = qs.filter(prod_nome__icontains=prod_nome)

        if prod_codi:
            qs = qs.filter(prod_codi__icontains=prod_codi)

        # Ordenação correta do código como número
        qs = qs.annotate(
            prod_codi_int=Cast('prod_codi', IntegerField())
        )

        # Preços
        preco_qs = Tabelaprecos.objects.using(self.db_alias).filter(
            tabe_prod=OuterRef('prod_codi')
        )

        if self.empresa_id:
            try:
                preco_qs = preco_qs.filter(tabe_empr=int(self.empresa_id))
            except Exception:
                pass

        if self.filial_id:
            try:
                preco_qs = preco_qs.filter(tabe_fili=int(self.filial_id))
            except Exception:
                pass

        qs = qs.annotate(
            preco_vista=Coalesce(
                Subquery(preco_qs.values('tabe_avis')[:1], output_field=DecimalField()),
                0
            ),
            preco_prazo=Coalesce(
                Subquery(preco_qs.values('tabe_praz')[:1], output_field=DecimalField()),
                0
            ),
            preco_custo=Coalesce(
                Subquery(preco_qs.values('tabe_cuge')[:1], output_field=DecimalField()),
                0
            ),
        ).order_by('prod_empr', 'prod_codi_int')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Produtos'

        headers = [
            'Empresa',
            'Código',
            'Nome',
            'Unidade',
            'Grupo',
            'Subgrupo',
            'Família',
            'Marca',
            'NCM',
            'Preço à Vista',
            'Preço a Prazo',
            'Preço de Custo',
        ]
        ws.append(headers)

        header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center')

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for p in qs:
            ws.append([
                p.prod_empr or '',
                p.prod_codi or '',
                p.prod_nome or '',
                getattr(p.prod_unme, 'unid_desc', '') if p.prod_unme else '',
                getattr(p.prod_grup, 'descricao', '') if p.prod_grup else '',
                getattr(p.prod_sugr, 'descricao', '') if p.prod_sugr else '',
                getattr(p.prod_fami, 'descricao', '') if p.prod_fami else '',
                getattr(p.prod_marc, 'nome', '') if p.prod_marc else '',
                p.prod_ncm or '',
                float(p.preco_vista or 0),
                float(p.preco_prazo or 0),
                float(p.preco_custo or 0),
            ])

        larguras = {
            'A': 10,
            'B': 12,
            'C': 40,
            'D': 18,
            'E': 22,
            'F': 22,
            'G': 22,
            'H': 22,
            'I': 16,
            'J': 16,
            'K': 16,
            'L': 16,
        }

        for col, largura in larguras.items():
            ws.column_dimensions[col].width = largura

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="produtos.xlsx"'

        wb.save(response)
        return response


class ProdutoFotoView(DBAndSlugMixin, View):
    def get(self, request, *args, **kwargs):
        prod_codi = kwargs.get('prod_codi')
        if not prod_codi:
            raise Http404('Código do produto não informado')
        qs = Produtos.objects.using(self.db_alias).filter(prod_codi=prod_codi)
        if self.empresa_id:
            qs = qs.filter(prod_empr=str(self.empresa_id))
        produto = qs.order_by('prod_empr').first()
        if not produto:
            raise Http404('Produto não encontrado')

        foto = produto.prod_foto
        if not foto:
            # Sem foto, retornar 404 para que o template use placeholder
            raise Http404('Foto não disponível')
        return HttpResponse(bytes(foto), content_type='image/jpeg')



class SaldosDashboardView(DBAndSlugMixin, TemplateView):
    template_name = 'Produtos/saldos.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['slug'] = self.slug
        banco = self.db_alias
        
        # Filtros (Prioridade: GET > Atributos da View > Sessão > Headers > Default)
        empresa = self.request.GET.get('empresa') or self.empresa_id or self.request.session.get('empresa_id') or self.request.headers.get('X-Empresa') or 1
        filial = self.request.GET.get('filial') or self.filial_id or self.request.session.get('filial_id') or self.request.headers.get('X-Filial') or 1

        # Filtros de produto e data
        raw_list = self.request.GET.getlist('produto')
        raw_str = (self.request.GET.get('produto') or '').strip()
        if raw_str and not raw_list:
            raw_list = [x.strip() for x in raw_str.split(',') if x.strip()]
        produtos_sel = [str(x) for x in raw_list if x]
        data_inicio = (self.request.GET.get('data_inicio') or '').strip()
        data_fim = (self.request.GET.get('data_fim') or '').strip()
        
        filtro_empresa = Empresas.objects.using(banco).filter(empr_codi=empresa).first()
        # Filtra filial considerando também a empresa para garantir consistência
        filtro_filial = Filiais.objects.using(banco).filter(empr_empr=filial).first()
        if filtro_filial and str(filtro_filial.empr_codi) != str(empresa):
             # Se a filial não pertencer à empresa selecionada, tenta pegar a primeira filial da empresa
             filtro_filial = Filiais.objects.using(banco).filter(empr_codi=empresa).first()
             if filtro_filial:
                 filial = filtro_filial.empr_empr

        # Carregar listas para os selects de filtro
        try:
            ctx['empresas'] = Empresas.objects.using(banco).all().order_by('empr_nome')
            
            filiais_qs = Filiais.objects.using(banco).all()
            if empresa:
                filiais_qs = filiais_qs.filter(empr_codi=empresa)
            ctx['filiais'] = filiais_qs.order_by('empr_nome')
        except Exception:
            ctx['empresas'] = []
            ctx['filiais'] = []    

        # Defaults de período: últimos 30 dias
        from datetime import date, timedelta
        if not data_inicio or not data_fim:
            hoje = date.today()
            inicio = hoje - timedelta(days=30)
            data_inicio = data_inicio or inicio.strftime('%Y-%m-%d')
            data_fim = data_fim or hoje.strftime('%Y-%m-%d')

        # Produtos para o select
        try:
            produtos_qs = Produtos.objects.using(banco).all()
            if empresa:
                produtos_qs = produtos_qs.filter(prod_empr=str(empresa))
            ctx['produtos'] = produtos_qs.order_by('prod_nome')[:300]
        except Exception:
            ctx['produtos'] = []

        # Entradas e Saídas agregadas por produto
        entradas_data = []
        saidas_data = []
        try:
            from Entradas_Estoque.models import EntradaEstoque
            from Saidas_Estoque.models import SaidasEstoque
            from datetime import datetime
            di = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            df = datetime.strptime(data_fim, '%Y-%m-%d').date()

            ent_qs = EntradaEstoque.objects.using(banco).filter(entr_empr=int(empresa), entr_fili=int(filial), entr_data__range=(di, df))
            sai_qs = SaidasEstoque.objects.using(banco).filter(said_empr=int(empresa), said_fili=int(filial), said_data__range=(di, df))
            if produtos_sel:
                ent_qs = ent_qs.filter(entr_prod__in=produtos_sel)
                sai_qs = sai_qs.filter(said_prod__in=produtos_sel)

            from django.db.models import Sum
            ent_group = ent_qs.values('entr_prod').annotate(total_entradas=Sum('entr_quan')).order_by('entr_prod')[:200]
            sai_group = sai_qs.values('said_prod').annotate(total_saidas=Sum('said_quan')).order_by('said_prod')[:200]
            entradas_data = list(ent_group)
            saidas_data = list(sai_group)
        except Exception:
            entradas_data = []
            saidas_data = []

        chart_labels = []
        chart_entradas = []
        chart_saidas = []
        try:
            cods_ent = [e.get('entr_prod') for e in entradas_data]
            cods_sai = [s.get('said_prod') for s in saidas_data]
            cods = sorted(set([c for c in cods_ent + cods_sai if c]))
            ent_map = {e.get('entr_prod'): float(e.get('total_entradas') or 0) for e in entradas_data}
            sai_map = {s.get('said_prod'): float(s.get('total_saidas') or 0) for s in saidas_data}
            nomes = {}
            if cods:
                prods_nomes = Produtos.objects.using(banco).filter(prod_codi__in=cods)
                nomes = {p.prod_codi: p.prod_nome for p in prods_nomes}
            for c in cods:
                chart_labels.append(nomes.get(c, c))
                chart_entradas.append(ent_map.get(c, 0))
                chart_saidas.append(sai_map.get(c, 0))
        except Exception:
            chart_labels = []
            chart_entradas = []
            chart_saidas = []

        try:
            combined = []
            for i, lbl in enumerate(chart_labels):
                ve = float(chart_entradas[i] if i < len(chart_entradas) else 0)
                vs = float(chart_saidas[i] if i < len(chart_saidas) else 0)
                combined.append((lbl, ve, vs, ve + vs))
            combined.sort(key=lambda x: x[3], reverse=True)
            combined = combined[:30]
            chart_labels = [c[0] for c in combined]
            chart_entradas = [c[1] for c in combined]
            chart_saidas = [c[2] for c in combined]
        except Exception:
            pass

        try:
            kpi_total_entradas = float(sum(chart_entradas))
            kpi_total_saidas = float(sum(chart_saidas))
            kpi_produtos_mov = int(len(chart_labels))
        except Exception:
            kpi_total_entradas = 0.0
            kpi_total_saidas = 0.0
            kpi_produtos_mov = 0

        # Saldos atuais por produto
        saldos_list = []
        try:
            from decimal import Decimal
            from django.db.models import Sum
            from Produtos.models import Lote
            if filtro_empresa:
                saldos_qs = SaldoProduto.objects.using(banco).filter(empresa=str(empresa))
                if filtro_filial:
                    saldos_qs = saldos_qs.filter(filial=str(filial))
            else:
                saldos_qs = SaldoProduto.objects.using(banco).filter(empresa__in=[str(empresa), '0'], filial=str(filial))
            if produtos_sel:
                saldos_qs = saldos_qs.filter(produto_codigo__in=produtos_sel)
            saldos_qs = saldos_qs.order_by('produto_codigo')[:300]
            codigos = [str(getattr(s, 'produto_codigo_id', None) or getattr(s, 'produto_codigo', '')) for s in saldos_qs]
            nomes_map = {}
            if codigos:
                prods = Produtos.objects.using(banco).filter(prod_codi__in=codigos)
                nomes_map = {p.prod_codi: p.prod_nome for p in prods}
            lotes_map = {}
            try:
                lotes_group = (
                    Lote.objects.using(banco)
                    .filter(lote_empr=int(empresa), lote_prod__in=codigos, lote_ativ=True)
                    .values('lote_prod')
                    .annotate(total=Sum('lote_sald'))
                )
                lotes_map = {str(r.get('lote_prod')): Decimal(str(r.get('total') or 0)) for r in lotes_group}
            except Exception:
                lotes_map = {}
            for s in saldos_qs:
                codigo = str(getattr(s, 'produto_codigo_id', None) or getattr(s, 'produto_codigo', ''))
                saldo_total = Decimal(str(getattr(s, 'saldo_estoque', 0) or 0))
                saldo_lotes = lotes_map.get(codigo, Decimal('0'))
                saldo_sem_lote = saldo_total - saldo_lotes
                saldos_list.append({
                    'prod_codi': codigo,
                    'prod_nome': nomes_map.get(codigo, codigo),
                    'saldo_total': saldo_total,
                    'saldo_lotes': saldo_lotes,
                    'saldo_sem_lote': saldo_sem_lote,
                })
        except Exception:
            saldos_list = []

        # Detalhe de saldo do produto selecionado
        saldo_prod_sel = None
        if produtos_sel:
            try:
                if filtro_empresa:
                    sp = SaldoProduto.objects.using(banco).filter(produto_codigo__in=produtos_sel, empresa=str(empresa), filial=str(filial)).first()
                else:
                    sp = SaldoProduto.objects.using(banco).filter(produto_codigo__in=produtos_sel, empresa=str(empresa), filial=str(filial)).first()
                if sp:
                    saldo_prod_sel = getattr(sp, 'saldo_estoque', None)
            except Exception:
                saldo_prod_sel = None

        try:
            empresa_val = int(empresa)
        except (ValueError, TypeError):
            empresa_val = empresa
            
        try:
            filial_val = int(filial)
        except (ValueError, TypeError):
            filial_val = filial

        filtros_ctx = {
            'empresa': empresa_val,
            'filial': filial_val,
            'data_inicial': data_inicio,
            'data_final': data_fim,
            'produto': ','.join(produtos_sel) if produtos_sel else '',
            'filtro_empresa': filtro_empresa,
            'filtro_filial': filtro_filial,
        }

        ctx.update({
            'entradas_data': entradas_data,
            'saidas_data': saidas_data,
            'saldos_produtos': saldos_list,
            'produtos_selecionados': produtos_sel,
            'saldo_produto_selecionado': saldo_prod_sel,
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'chart_labels': chart_labels,
            'chart_entradas': chart_entradas,
            'chart_saidas': chart_saidas,
            'filtros': filtros_ctx,
            'kpi_total_entradas': kpi_total_entradas,
            'kpi_total_saidas': kpi_total_saidas,
            'kpi_produtos_mov': kpi_produtos_mov,
        })
        return ctx


class SaldosMovimentosView(DBAndSlugMixin, View):
    def get(self, request, *args, **kwargs):
        banco = self.db_alias
        empresa = request.GET.get('empresa') or self.empresa_id or request.session.get('empresa_id') or request.headers.get('X-Empresa') or 1
        filial = request.GET.get('filial') or self.filial_id or request.session.get('filial_id') or request.headers.get('X-Filial') or 1
        produto = (request.GET.get('produto') or '').strip()
        data_inicio = (request.GET.get('data_inicio') or request.GET.get('data_inicial') or '').strip()
        data_fim = (request.GET.get('data_fim') or request.GET.get('data_final') or '').strip()
        limit = request.GET.get('limit')
        try:
            limit_i = int(limit) if limit is not None and str(limit).strip() != '' else 200
        except Exception:
            limit_i = 200

        from Produtos.consultas.estoque_consultas import obter_movimentacoes_produto

        payload = obter_movimentacoes_produto(
            banco=banco,
            empresa=empresa,
            filial=filial,
            produto=produto,
            data_inicio=data_inicio,
            data_fim=data_fim,
            limit=limit_i,
        )

        slug = self.slug or get_licenca_slug()
        for r in payload.get('entradas') or []:
            r['link'] = f"/web/{slug}/entradas/{r.get('entr_sequ')}/"
        for r in payload.get('saidas') or []:
            r['link'] = f"/web/{slug}/saidas/{r.get('said_sequ')}/"

        return JsonResponse(
            {
                'banco': banco,
                'slug': slug,
                'empresa': str(empresa),
                'filial': str(filial),
                'produto': produto,
                **payload,
            },
            json_dumps_params={'default': str},
        )

def autocomplete_produtos(request, slug=None):
    banco = get_licenca_db_config(request) or 'default'
    empresa_id = request.session.get('empresa_id', 1)
    term = (request.GET.get('term') or request.GET.get('q') or '').strip()
    qs = Produtos.objects.using(banco).filter(prod_empr=str(empresa_id))
    if term:
        if term.isdigit():
            qs = qs.filter(prod_codi__icontains=term)
        else:
            qs = qs.filter(prod_nome__icontains=term)
    qs = qs.order_by('prod_nome')[:20]
    data = [{'id': str(obj.prod_codi), 'text': f"{obj.prod_codi} - {obj.prod_nome}"} for obj in qs]
    return JsonResponse({'results': data})


class ZerarEstoqueView(DBAndSlugMixin, TemplateView):
    template_name = 'Produtos/zerar_estoque.html'

    def _get_contexto(self):
        empresa = self.request.GET.get('empresa') or self.empresa_id or self.request.session.get('empresa_id') or 1
        filial = self.request.GET.get('filial') or self.filial_id or self.request.session.get('filial_id') or 1
        return str(empresa), str(filial)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        empresa, filial = self._get_contexto()
        ctx['slug'] = self.slug
        ctx['empresa'] = empresa
        ctx['filial'] = filial

        try:
            from Produtos.servicos.zerar_estoque import obter_saldos_atuais

            info = obter_saldos_atuais(
                banco=self.db_alias,
                empresa=empresa,
                filial=filial,
                apenas_com_saldo=True,
                limit=50,
            )
            ctx['total_com_saldo'] = info.get('total', 0)
            ctx['amostra'] = info.get('amostra', [])
        except Exception:
            ctx['total_com_saldo'] = 0
            ctx['amostra'] = []

        return ctx

    def post(self, request, *args, **kwargs):
        empresa = request.POST.get('empresa') or self.empresa_id or request.session.get('empresa_id') or 1
        filial = request.POST.get('filial') or self.filial_id or request.session.get('filial_id') or 1
        limit_resultados = request.POST.get('limit_resultados') or 2000
        batch_size = request.POST.get('batch_size') or 500

        try:
            from Produtos.servicos.zerar_estoque import zerar_estoque

            resultado = zerar_estoque(
                banco=self.db_alias,
                empresa=empresa,
                filial=filial,
                batch_size=int(batch_size or 500),
                limit_resultados=None if str(limit_resultados).strip().lower() in ('none', '') else int(limit_resultados),
            )
            messages.success(request, f"Estoque zerado para empresa {resultado.get('empresa')} filial {resultado.get('filial')}. Itens afetados: {resultado.get('zerados')}.")
        except Exception as e:
            logger.exception(f"Falha ao zerar estoque: {e}")
            messages.error(request, f"Falha ao zerar estoque: {e}")
            resultado = None

        ctx = self.get_context_data(**kwargs)
        ctx['resultado'] = resultado
        return self.render_to_response(ctx)
